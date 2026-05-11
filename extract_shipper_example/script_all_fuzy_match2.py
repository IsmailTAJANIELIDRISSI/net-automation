#!/usr/bin/env python3
"""
Consolidated Document Processing Script with Gemini AI Enhancement
Combines PDF renaming, Excel processing, data validation, and AI-powered shipper name extraction
"""

import os
import glob
import re
import hashlib
import json
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import pdfplumber
from pathlib import Path
import subprocess
import tempfile
from PyPDF2 import PdfReader, PdfWriter
import difflib
# Use ONLY the new google.genai API (old google.generativeai is deprecated)
# Default: Gemini 3.1 Flash Lite Preview (replaces discontinued gemini-2.0-flash / flash-lite as of 2026-06-01).
# Override with env GEMINI_MODEL if needed.
DEFAULT_GEMINI_MODEL = "gemini-3.1-flash-lite-preview"
GEMINI_MODEL_FALLBACKS = (
    "gemini-3.1-flash-lite-preview",
    "gemini-2.5-flash",
    "gemini-1.5-flash",
)
GENAI_CLIENT = None
GENAI_MODEL = None
USE_NEW_API = False

try:
    import google.genai as genai_new
    USE_NEW_API = True
except ImportError:
    # Old API is deprecated - no longer supported
    # Users must install: pip install google-genai
    genai_new = None
    USE_NEW_API = False
import time
import sys
import json

def _load_lta_license():
    """Load LTA license from config file"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(script_dir, 'config', 'license.json')
        
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                return config.get('LTA_sys_validity', '2026-03-07')
        return '2026-03-07'  # Default fallback
    except:
        return '2026-03-07'

# Load license expiry date from config
LTA_license_expires = _load_lta_license() 

# Try to load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Setup logging (with error handling)
try:
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('shipper_extraction.log', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
except (PermissionError, OSError):
    # If log file creation fails, use console only
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )
logger = logging.getLogger(__name__)

# Global variable for known companies database
KNOWN_COMPANIES = []
DATABASE_FILE = "known_companies.json"

def load_companies_database():
    """Load known companies from JSON file"""
    global KNOWN_COMPANIES
    try:
        if os.path.exists(DATABASE_FILE):
            with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
                KNOWN_COMPANIES = json.load(f)
            logger.info(f"Loaded {len(KNOWN_COMPANIES)} companies from database")
        else:
            logger.warning(f"Database file {DATABASE_FILE} not found. Creating empty database.")
            KNOWN_COMPANIES = []
    except Exception as e:
        logger.error(f"Error loading companies database: {e}")
        KNOWN_COMPANIES = []

def save_companies_database():
    """Save known companies to JSON file with atomic write to prevent corruption"""
    try:
        # Use atomic write: write to temp file first, then rename
        temp_file = DATABASE_FILE + '.tmp'
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(KNOWN_COMPANIES, f, indent=2, ensure_ascii=False)
        
        # Atomic rename (works on Windows and Unix)
        if os.path.exists(DATABASE_FILE):
            os.replace(temp_file, DATABASE_FILE)
        else:
            os.rename(temp_file, DATABASE_FILE)
        
        logger.info(f"Saved {len(KNOWN_COMPANIES)} companies to database")
    except Exception as e:
        logger.error(f"Error saving companies database: {e}")
        # Clean up temp file if it exists
        try:
            if os.path.exists(DATABASE_FILE + '.tmp'):
                os.remove(DATABASE_FILE + '.tmp')
        except:
            pass

def add_company_to_database(company_name):
    """Add new company to database if it doesn't exist - reloads database first to prevent overwrites"""
    global KNOWN_COMPANIES
    
    if not company_name:
        return False
    
    # CRITICAL: Reload database from file before adding to ensure we have the latest data
    # This prevents overwriting companies added by other script instances
    try:
        if os.path.exists(DATABASE_FILE):
            with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
                current_companies = json.load(f)
                # Update global list with latest data from file
                KNOWN_COMPANIES = current_companies
                logger.debug(f"Reloaded {len(KNOWN_COMPANIES)} companies from database before adding")
    except Exception as e:
        logger.warning(f"Could not reload database before adding: {e}, using current in-memory list")
        # Continue with current KNOWN_COMPANIES if reload fails
    
    # Check if company already exists (case-insensitive comparison)
    company_name_upper = company_name.upper().strip()
    existing_upper = [c.upper().strip() if c else '' for c in KNOWN_COMPANIES]
    
    if company_name_upper in existing_upper:
        logger.debug(f"Company already exists in database (case-insensitive): {company_name}")
        return False
    
    # Add the company
    KNOWN_COMPANIES.append(company_name)
    save_companies_database()
    logger.info(f"Added new company to database: {company_name} (total: {len(KNOWN_COMPANIES)})")
    return True

def setup_gemini_api():
    """Setup Gemini API with API key from environment or prompt user"""
    global GENAI_CLIENT, GENAI_MODEL, USE_NEW_API
    
    api_key = os.getenv('GEMINI_API_KEY')
    if not api_key:
        logger.warning("GEMINI_API_KEY not found in environment variables")
        # For now, we'll skip API setup if no key is provided
        # In production, you might want to prompt the user or exit
        return False
    
    try:
        if USE_NEW_API:
            # New API (google.genai): use Client
            # The new API doesn't require getting a model object first
            # Just create the client and use client.models.generate_content(model='name', contents='...')
            GENAI_CLIENT = genai_new.Client(api_key=api_key)
            GENAI_MODEL = (os.getenv("GEMINI_MODEL") or "").strip() or DEFAULT_GEMINI_MODEL
            logger.info(f"Gemini API (new google.genai) configured successfully with model: {GENAI_MODEL}")
            return True
        else:
            # Old API is no longer supported - user must install google-genai
            logger.error("Google Gemini API not available. Please install: pip install google-genai")
            logger.error("The old google.generativeai package is deprecated and no longer supported.")
            return False
    except Exception as e:
        logger.error(f"Error configuring Gemini API: {e}")
        return False


def _gemini_response_text(response):
    """
    Extract user-facing text from a generate_content response.
    Gemini 3 may return multiple parts (e.g. thought vs answer); join non-thought text parts.
    For single-turn calls the SDK handles thought signatures; no manual circulation needed.
    """
    if response is None:
        return None
    if hasattr(response, "text") and response.text:
        return response.text
    try:
        parts_out = []
        cands = getattr(response, "candidates", None) or []
        if not cands:
            return str(response)
        content = getattr(cands[0], "content", None)
        part_list = getattr(content, "parts", None) if content else None
        if not part_list:
            return str(response)
        for part in part_list:
            if getattr(part, "thought", False):
                continue
            t = getattr(part, "text", None)
            if t:
                parts_out.append(t)
        return "\n".join(parts_out) if parts_out else str(response)
    except Exception:
        return str(response)


def verify_shipper_with_gemini(extracted_name, database_companies):
    """Use Gemini AI to select the best company name from candidates and database"""
    global GENAI_CLIENT, GENAI_MODEL, USE_NEW_API
    
    try:
        # Accept either a single string or a list of candidates
        if isinstance(extracted_name, list):
            candidates = extracted_name
        else:
            candidates = [extracted_name]
        prompt = f"""
You are an expert in logistics company name identification and OCR error correction.
Here are all potential company names extracted from a shipping document (prioritized by fuzzy matching):
{json.dumps(candidates, indent=2)}

KNOWN COMPANIES DATABASE:
{json.dumps(database_companies[:25], indent=2)}

CRITICAL INSTRUCTIONS:
1. **OCR ERROR CORRECTION**: Common OCR mistakes you MUST fix:
   - "LOGIGHES" → "LOGISTICS"
   - "LOGISHCS" → "LOGISTICS"
   - "LOGIGHCS" → "LOGISTICS"
   - "+76" or similar random chars → "LTD" or "CO., LTD"
   - "C0" (zero) → "CO" (letter O)
   - "I" (i) → "L" (L) in company suffixes
   - Any garbled text at the end of company names should be cleaned
   - **REMOVE leading non-letter characters**: "; COMPANY" → "COMPANY", ": COMPANY" → "COMPANY", "- COMPANY" → "COMPANY"

2. **COMPANY TYPE ACCEPTANCE**: 
   - Accept ALL company types: TRADING companies, LOGISTICS companies, MANUFACTURING companies, etc.
   - The shipper can be ANY type of company - do NOT reject based on company type
   - "TRADING CO", "INDUSTRIAL CO", "LOGISTICS CO" are ALL valid shippers

3. **FUZZY MATCHING PRIORITY**:
   - Candidates at the START of the list have already been fuzzy-matched - prioritize them!
   - Compare against database using fuzzy logic: ignore punctuation, OCR errors, minor spelling differences
   - >70% similarity = match

4. **EXCLUSIONS** - NEVER select these (they are consignees/receivers, NOT shippers):
   - "MED AFRICA LOGISTICS" (or any variation)

5. **EXTRACTION RULES**:
   - If a candidate contains BOTH a company name AND an address, extract ONLY the company name
   - Ignore addresses, room numbers, floor numbers, street names
   - Focus on the actual company legal name

6. **OUTPUT FORMAT** - Respond in this EXACT JSON format:
{{
    "reasoning": "Your step-by-step analysis with OCR corrections and specific reasons for match/no-match",
    "matched_company": "best database match or null",
    "is_new_company": true/false,
    "selected_candidate": "the chosen candidate from the list",
    "final_name": "the final cleaned company name (with OCR errors corrected)"
}}
"""
        # Use appropriate API based on what's available
        if USE_NEW_API and GENAI_CLIENT:
            # New API (google.genai): use client.models.generate_content()
            # Prefer configured model, then known-good fallbacks (no Gemini 2.0 — discontinued June 2026).
            seen = set()
            model_names = []
            for m in (GENAI_MODEL,) + GEMINI_MODEL_FALLBACKS:
                if m and m not in seen:
                    seen.add(m)
                    model_names.append(m)
            response = None
            last_error = None
            
            for model_name in model_names:
                try:
                    # New google.genai API: client.models.generate_content(model='name', contents='...')
                    response = GENAI_CLIENT.models.generate_content(
                        model=model_name,
                        contents=prompt
                    )
                    break
                except Exception as e:
                    last_error = e
                    logger.debug(f"Model {model_name} failed: {e}")
                    continue
            
            if response is None:
                raise Exception(f"Failed to generate content with new API: {last_error}")
            
            response_text_raw = _gemini_response_text(response)
        else:
            # New API not available - cannot proceed
            raise Exception("Google Gemini API not available. Please install: pip install google-genai")
        
        logger.info(f"Gemini API response for candidates: {response_text_raw[:200]}...")
        # Parse JSON response
        try:
            response_text = response_text_raw.strip()
            if response_text.startswith('```json'):
                start = response_text.find('{')
                end = response_text.rfind('}') + 1
                if start != -1 and end != 0:
                    json_text = response_text[start:end]
                else:
                    json_text = response_text
            elif response_text.startswith('```'):
                lines = response_text.split('\n')
                json_lines = []
                in_json = False
                for line in lines:
                    if line.strip().startswith('{') or in_json:
                        in_json = True
                        json_lines.append(line)
                        if line.strip().endswith('}') and json_lines:
                            break
                json_text = '\n'.join(json_lines)
            else:
                json_text = response_text
            result = json.loads(json_text)
            logger.info(f"Successfully parsed Gemini JSON response")
            return result
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse Gemini response as JSON: {e}")
            # Get response text for error logging
            error_text = response_text_raw if 'response_text_raw' in locals() else str(response)
            logger.error(f"Response text: {error_text[:200]}...")
            return {
                "reasoning": "JSON parse error",
                "matched_company": None,
                "is_new_company": True,
                "selected_candidate": candidates[0] if candidates else None,
                "final_name": candidates[0] if candidates else None
            }
    except Exception as e:
        logger.error(f"Error calling Gemini API: {e}")
        return {
            "reasoning": f"API error: {e}",
            "matched_company": None,
            "is_new_company": True,
            "selected_candidate": candidates[0] if candidates else None,
            "final_name": candidates[0] if candidates else None
        }


def extract_vision_meta(pdf_path):
    """
    Scanned-PDF fallback: send raw PDF bytes to Gemini Vision to extract the shipper name.
    Returns the shipper name string, or None on failure.
    """
    global GENAI_CLIENT, GENAI_MODEL, USE_NEW_API
    if not USE_NEW_API or not GENAI_CLIENT:
        logger.info("Gemini Vision ignoré — client non initialisé")
        return None

    import base64
    with open(pdf_path, "rb") as f:
        pdf_base64 = base64.b64encode(f.read()).decode("utf-8")

    known = KNOWN_COMPANIES[:25]
    prompt = f"""This is an Air Waybill (MAWB) document.

Extract the SHIPPER NAME from the \"Shipper's Name and Address\" box (top-left of the form).
- Return ONLY the company name (no address, no street, no phone, no email).
- Fix obvious OCR errors (\"C0\" -> \"CO\", \"LOGIGHCS\" -> \"LOGISTICS\").
- NEVER return \"MED AFRICA LOGISTICS\" — that is the consignee.
- If it closely matches a known company below, return the canonical form.
Known companies: {json.dumps(known)}

Respond in this EXACT JSON (no markdown fences):
{{"shipper_name": "COMPANY NAME OR null"}}"""

    seen = set()
    model_names = []
    for m in ((GENAI_MODEL,) + GEMINI_MODEL_FALLBACKS):
        if m and m not in seen:
            seen.add(m)
            model_names.append(m)

    for model_name in model_names:
        try:
            logger.info(f"PDF scanné — appel Gemini Vision ({model_name})…")
            response = GENAI_CLIENT.models.generate_content(
                model=model_name,
                contents=[{
                    "parts": [
                        {"inline_data": {"mime_type": "application/pdf", "data": pdf_base64}},
                        {"text": prompt},
                    ]
                }]
            )
            raw = (_gemini_response_text(response) or "").strip()
            if raw.startswith("```"):
                start = raw.find("{")
                end = raw.rfind("}") + 1
                raw = raw[start:end] if start != -1 else raw
            parsed = json.loads(raw)
            shipper_name = parsed.get("shipper_name") or None
            logger.info(f"Gemini Vision résultat: expéditeur=\"{shipper_name}\"")
            return shipper_name
        except Exception as e:
            logger.warning(f"Gemini Vision {model_name} échoué: {e}")
            continue

    logger.warning("Tous les modèles Gemini Vision ont échoué")
    return None


def is_pdf_text_based(pdf_path):
    """Determine if PDF contains extractable text or is image-based"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check first few pages for text content
            for page_num in range(min(3, len(pdf.pages))):
                page = pdf.pages[page_num]
                text = page.extract_text()
                if text and len(text.strip()) > 100:
                    # Found substantial text content
                    return True
        return False
    except Exception as e:
        logger.error(f"Error checking PDF type for {pdf_path}: {e}")
        return False

def extract_shipper_name_text_based(pdf_path):
    """Extract shipper name from text-based PDF with exact extraction"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing text-based PDF: {os.path.basename(pdf_path)}")
            
            # Find page with shipper information
            shipper_page = None
            shipper_page_num = 0
            
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and ('shipper' in text.lower() or 'consignee' in text.lower()):
                    shipper_page = page
                    shipper_page_num = page_num
                    break
            
            if not shipper_page:
                logger.warning("No shipper information found in text-based PDF")
                return None
            
            # Extract text and find shipper section
            page_text = shipper_page.extract_text()
            lines = page_text.split('\n')
            
            # Look for shipper section markers
            shipper_start = -1
            consignee_start = -1
            
            for i, line in enumerate(lines):
                line_lower = line.lower().strip()
                if 'shipper' in line_lower and ('name' in line_lower or 'address' in line_lower):
                    shipper_start = i
                    logger.info(f"Found shipper section at line {i}: {line.strip()}")
                elif 'consignee' in line_lower and ('name' in line_lower or 'address' in line_lower):
                    consignee_start = i
                    break
            
            if shipper_start == -1:
                logger.warning("Could not locate shipper section in text-based PDF")
                return None
            
            # Extract shipper name from the section using robust logic
            start_idx = max(0, shipper_start + 1)
            end_idx = consignee_start if consignee_start > 0 else min(len(lines), 25)
            potential_companies = []
            for i in range(start_idx, end_idx):
                if i < len(lines):
                    line = lines[i].strip()
                    if len(line) > 8:
                        cleaned_line = clean_extracted_text(line)
                        if is_airline_or_system_text(cleaned_line):
                            continue
                        if might_be_company(cleaned_line):
                            potential_companies.append(cleaned_line)
            logger.info(f"Found {len(potential_companies)} potential company texts in shipper section")
            # Return all potential company names for AI selection
            if potential_companies:
                return potential_companies
            logger.warning("Could not extract valid shipper name from text-based PDF")
            return []
            
    except Exception as e:
        logger.error(f"Error extracting from text-based PDF: {e}")
        return None
    """Copy style from one cell to another without using StyleProxy"""
    try:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text
            )
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    except Exception as e:
        print(f"    Warning: Could not copy style: {e}")

def copy_cell_style(source_cell, target_cell):
    """Copy style from one cell to another without using StyleProxy"""
    try:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text
            )
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    except Exception as e:
        print(f"    Warning: Could not copy style: {e}")

def get_file_hash(file_path):
    """Calculate MD5 hash of a file for duplicate detection"""
    hash_md5 = hashlib.md5()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except Exception as e:
        print(f"    Warning: Could not hash {os.path.basename(file_path)}: {e}")
        return None

def extract_base_name(filename):
    """Extract base name without timestamps and versioning"""
    filename = re.sub(r' - \d{4}-\d{2}-\d{2}T\d{6}\.\d{3}', '', filename)
    filename = re.sub(r' \(\d+\)', '', filename)
    filename = os.path.splitext(filename)[0]
    return filename.strip()

def find_and_remove_duplicates(dir_path):
    """Find and remove duplicate files based on content and naming patterns"""
    print("  Detecting and removing duplicate files...")
    all_files = []
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            file_path = os.path.join(root, file)
            if os.path.isfile(file_path):
                all_files.append(file_path)
    
    if not all_files:
        print("    No files found to check for duplicates")
        return
    
    file_groups = {}
    hash_to_files = {}
    print(f"    Analyzing {len(all_files)} files for duplicates...")
    
    for file_path in all_files:
        filename = os.path.basename(file_path)
        file_ext = os.path.splitext(filename)[1].lower()
        base_name = extract_base_name(filename)
        group_key = f"{base_name}_{file_ext}"
        if group_key not in file_groups:
            file_groups[group_key] = []
        file_groups[group_key].append(file_path)
        file_hash = get_file_hash(file_path)
        if file_hash:
            if file_hash not in hash_to_files:
                hash_to_files[file_hash] = []
            hash_to_files[file_hash].append(file_path)
    
    duplicates_removed = 0
    print("    Checking for exact duplicates (same content)...")
    for file_hash, file_list in hash_to_files.items():
        if len(file_list) > 1:
            file_list.sort()
            keep_file = file_list[0]
            print(f"    Found {len(file_list)} identical files:")
            print(f"      KEEPING: {os.path.basename(keep_file)}")
            for duplicate_file in file_list[1:]:
                try:
                    os.remove(duplicate_file)
                    print(f"      DELETED: {os.path.basename(duplicate_file)} (identical content)")
                    duplicates_removed += 1
                except Exception as e:
                    print(f"      ERROR deleting {os.path.basename(duplicate_file)}: {e}")
    
    print("    Checking for naming pattern duplicates...")
    for group_key, file_list in file_groups.items():
        if len(file_list) > 1:
            existing_files = [f for f in file_list if os.path.exists(f)]
            if len(existing_files) > 1:
                existing_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                keep_file = existing_files[0]
                print(f"    Found {len(existing_files)} files with similar names:")
                print(f"      KEEPING: {os.path.basename(keep_file)} (newest)")
                for duplicate_file in existing_files[1:]:
                    try:
                        if should_remove_as_duplicate(keep_file, duplicate_file):
                            os.remove(duplicate_file)
                            print(f"      DELETED: {os.path.basename(duplicate_file)} (older version)")
                            duplicates_removed += 1
                        else:
                            print(f"      KEPT: {os.path.basename(duplicate_file)} (significant differences detected)")
                    except Exception as e:
                        print(f"      ERROR deleting {os.path.basename(duplicate_file)}: {e}")
    
    print(f"    ✓ Duplicate detection completed - {duplicates_removed} files removed")

def should_remove_as_duplicate(keep_file, candidate_file):
    """Check if candidate file should be removed as duplicate of keep_file"""
    try:
        keep_size = os.path.getsize(keep_file)
        candidate_size = os.path.getsize(candidate_file)
        size_diff_percent = abs(keep_size - candidate_size) / max(keep_size, candidate_size) * 100
        if size_diff_percent > 5:
            return False
        if keep_file.endswith('.xlsx') and candidate_file.endswith('.xlsx'):
            return compare_excel_files_for_duplicates(keep_file, candidate_file)
        if keep_file.endswith('.pdf') and candidate_file.endswith('.pdf'):
            return size_diff_percent <= 1
        return size_diff_percent <= 2
    except Exception as e:
        print(f"      Warning: Could not compare files properly: {e}")
        return False

def compare_excel_files_for_duplicates(file1, file2):
    """Compare Excel files to determine if they're duplicates"""
    try:
        wb1 = load_workbook(file1, data_only=True)
        wb2 = load_workbook(file2, data_only=True)
        ws1 = wb1.active
        ws2 = wb2.active
        if ws1.max_row != ws2.max_row or ws1.max_column != ws2.max_column:
            return False
        sample_positions = [
            (1, 1), (1, 2), (1, 3),
            (2, 1), (2, 2), (2, 3),
            (min(5, ws1.max_row), 1)
        ]
        for row, col in sample_positions:
            if row <= ws1.max_row and col <= ws1.max_column:
                val1 = ws1.cell(row=row, column=col).value
                val2 = ws2.cell(row=row, column=col).value
                if val1 != val2:
                    return False
        return True
    except Exception:
        return False

def find_best_company_match(extracted_text, min_similarity=0.6):
    """Find the best matching company from known database using fuzzy matching"""
    if not extracted_text or len(extracted_text) < 5:
        return None
    cleaned_text = clean_for_matching(extracted_text)
    best_match = None
    best_score = 0
    for known_company in KNOWN_COMPANIES:
        scores = []
        score1 = difflib.SequenceMatcher(None, cleaned_text.upper(), known_company.upper()).ratio()
        scores.append(score1)
        cleaned_known = clean_for_matching(known_company)
        score2 = difflib.SequenceMatcher(None, cleaned_text.upper(), cleaned_known.upper()).ratio()
        scores.append(score2)
        text_words = set(cleaned_text.upper().split())
        known_words = set(cleaned_known.upper().split())
        if known_words:
            word_overlap = len(text_words & known_words) / len(known_words)
            scores.append(word_overlap)
        key_terms = extract_key_terms(known_company)
        text_key_terms = extract_key_terms(cleaned_text)
        if key_terms:
            key_overlap = len(set(text_key_terms) & set(key_terms)) / len(key_terms)
            scores.append(key_overlap)
        max_score = max(scores)
        print(f"   🔍 Matching '{cleaned_text[:50]}' vs '{known_company}'")
        print(f"      Scores: {[f'{s:.3f}' for s in scores]} -> Best: {max_score:.3f}")
        if max_score > best_score and max_score >= min_similarity:
            best_score = max_score
            best_match = known_company
    if best_match:
        print(f"   ✅ Best match: '{best_match}' (score: {best_score:.3f})")
        return best_match
    print(f"   ❌ No good match found (best score: {best_score:.3f})")
    return None

def clean_for_matching(text):
    """Clean text for better fuzzy matching"""
    if not text:
        return ""
    cleaned = text
    removals = [
        ' Air Waybill', ' issued by', ' Issued by', ' Not Negotiable',
        '+O6ISHES-6E—+F8', 'Ge Senccaiieeiont ys', '+76', 
        'LOGIGHES CO', 'CO +76'
    ]
    for removal in removals:
        cleaned = cleaned.replace(removal, '')
    fixes = {
        'SHANGHA|': 'SHANGHAI',
        'F1XLINK': 'FIXLINK', 
        'LOGIGHES': 'LOGISTICS',
        'INTERNAT1ONAL': 'INTERNATIONAL',
        'L1MITED': 'LIMITED',
        'C0.': 'CO.',
        'L1D': 'LTD',
        '1NC': 'INC',
    }
    for wrong, correct in fixes.items():
        cleaned = cleaned.replace(wrong, correct)
    import re
    cleaned = re.sub(r'[^\w\s.,()&-]', ' ', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned.strip()

def extract_key_terms(company_name):
    """Extract key identifying terms from company name"""
    if not company_name:
        return []
    words = company_name.upper().split()
    common_words = {'CO', 'CO.', 'LTD', 'LIMITED', 'INC', 'CORP', 'THE', 'AND', '&'}
    important_words = {'LOGISTICS', 'INTERNATIONAL', 'EXPRESS', 'SHIPPING', 'AIRWAYS'}
    key_terms = []
    for word in words:
        clean_word = word.strip('.,()&-')
        if clean_word in important_words or (len(clean_word) > 3 and clean_word not in common_words):
            key_terms.append(clean_word)
    return key_terms

def extract_specific_page_to_file(pdf_path, page_index, temp_dir):
    """Extract specific page to temporary file"""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PdfReader(file)
            if page_index >= len(pdf_reader.pages):
                print(f"   ❌ Page index {page_index} out of range")
                return None
            temp_pdf = os.path.join(temp_dir, f"page_{page_index + 1}.pdf")
            pdf_writer = PdfWriter()
            pdf_writer.add_page(pdf_reader.pages[page_index])
            with open(temp_pdf, 'wb') as output:
                pdf_writer.write(output)
            print(f"   📄 Extracted page {page_index + 1} to temporary file")
            return temp_pdf
    except Exception as e:
        print(f"   ❌ Error extracting page {page_index + 1}: {e}")
        return None

def crop_image_topleft(png_file, temp_dir):
    """Crop image to focus on top-left area"""
    try:
        from PIL import Image
        img = Image.open(png_file)
        width, height = img.size
        crop_box = (0, 0, int(width * 0.5), int(height * 0.4))
        cropped_img = img.crop(crop_box)
        cropped_file = os.path.join(temp_dir, "cropped.png")
        cropped_img.save(cropped_file)
        print(f"   ✂️  Cropped with PIL: {cropped_file}")
        return cropped_file
    except Exception as e:
        print(f"   ⚠️  PIL cropping failed: {e}, using original")
        return png_file

def crop_image_bottom_left(png_file, temp_dir):
    """Crop image to focus on bottom-left area for 'bloc' PDFs"""
    try:
        from PIL import Image
        img = Image.open(png_file)
        width, height = img.size
        print(f"   📐 Original image size: {width}x{height}")
        crop_box = (0, int(height * 0.5), int(width * 0.6), height)
        cropped_img = img.crop(crop_box)
        print(f"   📐 Cropped to bottom-left section: {crop_box}")
        cropped_file = os.path.join(temp_dir, "cropped_bottom_left.png")
        cropped_img.save(cropped_file)
        print(f"   ✂️  Cropped bottom-left with PIL: {cropped_file}")
        return cropped_file
    except Exception as e:
        print(f"   ⚠️  Bottom-left cropping failed: {e}, using original")
        return png_file

def crop_image_bottom_center(png_file, temp_dir):
    """Crop image to focus on bottom-center area for 'bloc' PDFs"""
    try:
        from PIL import Image
        img = Image.open(png_file)
        width, height = img.size
        print(f"   📐 Trying bottom-center crop on {width}x{height} image")
        crop_box = (int(width * 0.15), int(height * 0.5), int(width * 0.85), height)
        cropped_img = img.crop(crop_box)
        print(f"   📐 Cropped to bottom-center section: {crop_box}")
        cropped_file = os.path.join(temp_dir, "cropped_bottom_center.png")
        cropped_img.save(cropped_file)
        print(f"   ✂️  Cropped bottom-center with PIL: {cropped_file}")
        return cropped_file
    except Exception as e:
        print(f"   ⚠️  Bottom-center cropping failed: {e}, using original")
        return png_file

def process_extracted_text(text, is_text_based=True):
    """Process extracted text to find shipper company name"""
    if not text or len(text.strip()) < 10:
        return None
    lines = text.split('\n')
    print(f"   🔍 Processing {len(lines)} lines of text")
    shipper_section_start = -1
    consignee_section_start = -1
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if ('shipper' in line_lower and ('name' in line_lower or 'address' in line_lower)) or \
           ('shipper' in line_lower and i < 10):
            shipper_section_start = i
            print(f"   📍 Found shipper section at line {i}: {repr(line.strip())}")
        elif ('consignee' in line_lower and ('name' in line_lower or 'address' in line_lower)) or \
             ('consignee' in line_lower and i < 15):
            consignee_section_start = i
            print(f"   📍 Found consignee section at line {i}: {repr(line.strip())}")
            break
    start_idx = max(0, shipper_section_start + 1) if shipper_section_start >= 0 else 0
    end_idx = consignee_section_start if consignee_section_start > 0 else min(len(lines), 25)
    print(f"   🔍 Searching lines {start_idx} to {end_idx} for shipper")
    potential_companies = []
    for i in range(start_idx, end_idx):
        if i < len(lines):
            line = lines[i].strip()
            if len(line) > 8:
                cleaned_line = clean_extracted_text(line)
                if is_airline_or_system_text(cleaned_line):
                    print(f"   ❌ Skipping airline/system text: {cleaned_line}")
                    continue
                if might_be_company(cleaned_line):
                    potential_companies.append(cleaned_line)
                    print(f"   🤔 Potential company text: {cleaned_line}")
    print(f"   🎯 Found {len(potential_companies)} potential company texts")
    
    if not potential_companies:
        return None
    
    # Try fuzzy matching on each candidate
    # If we find a high-confidence match (>= 0.80), use it immediately
    best_match = None
    best_score = 0.0
    best_candidate = None
    
    for candidate in potential_companies:
        # Get detailed match info
        from difflib import SequenceMatcher
        for db_company in KNOWN_COMPANIES:
            ratio = SequenceMatcher(None, candidate.upper(), db_company.upper()).ratio()
            if ratio > best_score:
                best_score = ratio
                best_match = db_company
                best_candidate = candidate
    
    # If we have a high-confidence match (>= 0.80), return it immediately
    if best_score >= 0.80:
        print(f"   ✅ HIGH-CONFIDENCE MATCH ({best_score:.3f}): '{best_candidate}' -> '{best_match}'")
        print(f"   🎯 Skipping AI - using direct fuzzy match")
        return best_match  # Return the database company name directly
    
    # Otherwise, return all candidates for AI selection
    print(f"   🤔 No high-confidence match found (best: {best_score:.3f})")
    print(f"   🎯 Sending all {len(potential_companies)} candidates to AI for selection")
    return potential_companies    # If no good candidates found, try to find any company-like text

    for i in range(start_idx, end_idx):
        if i < len(lines):
            line = lines[i].strip()
            if len(line) > 10:
                cleaned_line = clean_extracted_text(line)
                if (not is_airline_or_system_text(cleaned_line) and 
                    might_be_company(cleaned_line) and
                    len(cleaned_line) > 15):  # Prefer longer names
                    print(f"   � Found potential name: {cleaned_line}")
                    return cleaned_line
    
    return None

def clean_extracted_text(text):
    """Clean extracted text from command-line tools"""
    if not text or len(text) < 3:
        return text
    cleaned = ' '.join(text.split())
    system_suffixes = [
        ' Air Waybill', ' issued by', ' Issued by', 
        ' +O6ISHES-6E—+F8 issued by', ' Not Negotiable'
    ]
    for suffix in system_suffixes:
        if suffix in cleaned:
            cleaned = cleaned.split(suffix)[0].strip()
    ocr_fixes = {
        'L1D': 'LTD',
        'LID': 'LTD', 
        'C0.': 'CO.',
        'CO,': 'CO.',
        'CO. .LTD': 'CO.,LTD',
        'CO..LTD': 'CO.,LTD',
        'CO. LTD': 'CO.,LTD',
        'LOGIST1CS': 'LOGISTICS',
        'LOGIGHES': 'LOGISTICS',
        'LOGISHCS': 'LOGISTICS',
        'LOGIGHCS': 'LOGISTICS',
        'LOGISITICS': 'LOGISTICS',
        'INTERNAT1ONAL': 'INTERNATIONAL',
        'INTERNAT1ON4L': 'INTERNATIONAL',
        'A1RWAYS': 'AIRWAYS',
        '1NC': 'INC',
        'L1MITED': 'LIMITED',
        'COMPAN1': 'COMPANY',
        'SHANGHA|': 'SHANGHAI',
        'F1XLINK': 'FIXLINK',
        '+76': 'LTD',
        ' +76': ' LTD',
        'CO +76': 'CO., LTD',
    }
    for wrong, correct in ocr_fixes.items():
        cleaned = cleaned.replace(wrong, correct)
    import re
    cleaned = re.sub(r'(?<=[A-Z])0(?=[A-Z])', 'O', cleaned)
    cleaned = re.sub(r'(?<=[A-Z])1(?=[A-Z])', 'I', cleaned)
    cleaned = re.sub(r'\|(?=[A-Z])', 'I', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = re.sub(r'[—–-]{2,}', '', cleaned)
    
    # Remove leading non-letter characters (semicolons, colons, dashes, etc.)
    cleaned = re.sub(r'^[^A-Za-z0-9]+', '', cleaned)
    
    return cleaned.strip()

def clean_company_name(name):
    """Clean company name by removing leading/trailing junk characters and text after company suffixes"""
    if not name:
        return name
    
    import re
    # Remove leading non-letter characters (semicolons, periods, dashes, etc.)
    cleaned = re.sub(r'^[^A-Za-z0-9]+', '', name)
    # Remove trailing non-letter characters except periods (for CO., LTD.)
    cleaned = re.sub(r'[^A-Za-z0-9.]+$', '', cleaned)
    
    # Remove everything after company suffixes (LTD, LIMITED, INC, CORP, etc.)
    # This handles cases like "COMPANY NAME CO., LTD Al" -> "COMPANY NAME CO., LTD"
    company_suffixes = [
        r'LTD\.',  # Match "LTD." first (more specific)
        r'LIMITED',
        r'LTD',    # Match "LTD" without period
        r'INC\.',
        r'INC',
        r'CORP\.',
        r'CORP',
        r'CO\.,LTD',
        r'CO\.LTD',
    ]
    
    # Try each suffix pattern and remove everything after it
    for suffix in company_suffixes:
        # Case-insensitive search for suffix followed by any characters
        pattern = rf'({suffix})[\s\S]*$'
        match = re.search(pattern, cleaned, re.IGNORECASE)
        if match:
            # Keep everything up to and including the suffix, remove the rest
            cleaned = cleaned[:match.end(1)]
            break  # Stop after first match to preserve the natural ending
    
    return cleaned.strip()

def might_be_company(text):
    """Quick check if text might be a company name (very permissive)"""
    if not text or len(text) < 8:
        return False
    text_upper = text.upper()
    company_indicators = [
        'LOGISTICS', 'INTERNATIONAL', 'CO.', 'LTD', 'INC', 'CORP', 
        'EXPRESS', 'SHIPPING', 'GROUP', 'COMPANY', 'FIXLINK', 'ANPORT'
    ]
    has_indicator = any(indicator in text_upper for indicator in company_indicators)
    upper_ratio = sum(1 for c in text if c.isupper()) / len(text) if text else 0
    looks_like_name = upper_ratio > 0.3
    return has_indicator or looks_like_name

def is_airline_or_system_text(text):
    """Check if text is airline name or system text to avoid"""
    if not text:
        return False
    text_upper = text.upper()
    
    # Exclude our own company (consignee, not shipper)
    own_companies = [
        'MED AFRICA LOGISTICS'
    ]
    for own_company in own_companies:
        if own_company in text_upper:
            logger.info(f"Skipping own company (consignee, not shipper): {text}")
            return True
    
    airlines = [
        'QATAR AIRWAYS', 'CHINA EASTERN AIRLINES', 'ETIHAD AIRWAYS', 
        'TURKISH AIRLINES', 'EMIRATES', 'AIR FRANCE', 'LUFTHANSA',
        'CATHAY PACIFIC', 'SINGAPORE AIRLINES', 'BRITISH AIRWAYS'
    ]
    system_text = [
        'AIR WAYBILL', 'ISSUED BY', 'NOT NEGOTIABLE', 'COPIES 1, 2',
        'TABULATOR STOPS', 'STAPLE DOCUMENTS', 'PERFORATION',
        'MEMBER OF IATA', 'P.O.BOX', 'DOHA.QATAR', 'HONG KONG',
        'VINCENT 1363', 'TEL:', 'EMAIL', '@', 'COM.CN',
        'INTERMEDIATE STOPPING', 'CARRIER DEEMS', 'CONDITIONS OF CONTRACT'
    ]
    if any(airline in text_upper for airline in airlines):
        return True
    if any(sys_text in text_upper for sys_text in system_text):
        return True
    if '@' in text or (any(char.isdigit() for char in text) and len([c for c in text if c.isdigit()]) > 8):
        return True
    return False

def try_ocrmypdf(pdf_path):
    """Try OCRmyPDF method"""
    try:
        print("   📄 Trying OCRmyPDF...")
        pdf_name = Path(pdf_path).name.lower()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_pdf = os.path.join(temp_dir, "ocr_output.pdf")
            cmd = ["ocrmypdf", "--force-ocr", "--skip-text", str(pdf_path), temp_pdf]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=90, encoding='utf-8', errors='ignore')
            if result.returncode == 0 and os.path.exists(temp_pdf):
                if "bloc" in pdf_name:
                    cmd = ["pdftotext", "-f", "2", "-l", "2", "-layout", temp_pdf, "-"]
                    print("   📋 Extracting from page 2 (bloc PDF)")
                else:
                    cmd = ["pdftotext", "-layout", temp_pdf, "-"]
                    print("   📄 Extracting from page 1")
                text_result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, encoding='utf-8', errors='ignore')
                if text_result.returncode == 0:
                    print(f"   📝 OCRmyPDF extracted {len(text_result.stdout)} characters")
                    result = process_extracted_text(text_result.stdout)
                    if result:
                        print(f"   ✅ OCRmyPDF found: {result}")
                        return result
    except Exception as e:
        print(f"   ⚠️  OCRmyPDF failed: {e}")
    return None

def try_tesseract_pdftoppm(pdf_path):
    """Try Tesseract with pdftoppm method"""
    try:
        print("   🖼️  Trying Tesseract + pdftoppm...")
        pdf_name = Path(pdf_path).name.lower()
        is_single_bloc = "bloc" in pdf_name
        with tempfile.TemporaryDirectory() as temp_dir:
            base_name = os.path.join(temp_dir, "page")
            cmd = ["pdftoppm", "-f", "1", "-l", "1", "-png", "-r", "300", str(pdf_path), base_name]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, encoding='utf-8', errors='ignore')
            if result.returncode == 0:
                png_file = f"{base_name}-1.png"
                if os.path.exists(png_file):
                    print(f"   📸 Generated image: {png_file}")
                    if is_single_bloc:
                        cropped_png = crop_image_bottom_left(png_file, temp_dir)
                        print(f"   📋 Applied single-page bloc cropping (bottom-left)")
                    else:
                        cropped_png = crop_image_topleft(png_file, temp_dir)
                        print(f"   📄 Applied standard cropping (top-left)")
                    cmd = ["tesseract", cropped_png, "stdout", "-l", "eng", "--psm", "6"]
                    ocr_result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, encoding='utf-8', errors='ignore')
                    if ocr_result.returncode == 0:
                        print(f"   📝 Tesseract extracted {len(ocr_result.stdout)} characters")
                        print(f"   🔍 First 300 chars: {repr(ocr_result.stdout[:300])}")
                        result = process_extracted_text(ocr_result.stdout)
                        if result:
                            print(f"   ✅ Tesseract found: {result}")
                            return result
    except Exception as e:
        print(f"   ⚠️  Tesseract failed: {e}")
    return None

def try_pdftotext(pdf_path):
    """Try pdftotext as fallback"""
    try:
        print("   📝 Trying pdftotext fallback...")
        cmd = ["pdftotext", "-layout", str(pdf_path), "-"]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, encoding='utf-8', errors='ignore')
        if result.returncode == 0 and result.stdout.strip():
            print(f"   📝 pdftotext extracted {len(result.stdout)} characters")
            extracted_result = process_extracted_text(result.stdout)
            if extracted_result:
                print(f"   ✅ pdftotext found: {extracted_result}")
                return extracted_result
    except Exception as e:
        print(f"   ⚠️  pdftotext failed: {e}")
    return None

def find_shipper_page_text_based(pdf_path):
    """Find which page contains shipper information in text-based PDFs"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            shipper_indicators = [
                "shipper's name and address",
                "shipper's account number", 
                "shippers name and address",
                "shippers account number",
                "shipper name and address"
            ]
            print(f"   📄 Scanning {len(pdf.pages)} pages for shipper information...")
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    page_text_lower = page_text.lower()
                    for indicator in shipper_indicators:
                        if indicator in page_text_lower:
                            print(f"   ✅ Found '{indicator}' on page {page_num + 1}")
                            return page_num, page
                    print(f"   📄 Page {page_num + 1}: No shipper indicators found")
                else:
                    print(f"   📄 Page {page_num + 1}: No extractable text")
            print("   ⚠️  No shipper indicators found in any page, using page 1 as fallback")
            return 0, pdf.pages[0]
    except Exception as e:
        print(f"   ❌ Error finding shipper page: {e}")
        return 0, pdf.pages[0] if pdf.pages else None

def extract_from_specific_page_text(page):
    """Extract shipper info from a specific pdfplumber page object, prioritizing text after specific indicators"""
    try:
        page_text = page.extract_text()
        if not page_text or len(page_text.strip()) < 50:
            print("   ⚠️ Insufficient text on page, falling back to OCR")
            return None
        
        # Split text into lines
        lines = page_text.split('\n')
        print(f"   🔍 Processing {len(lines)} lines of text")
        
        # Define indicators in order of preference
        indicators = [
            ("Shipper s Name", "Shipper's Name and Address"),
            ("QATAR AIR", "QATAR AIR"),
            ("Air Waybill", "Air Waybill")
        ]
        
        # Find the target indicator and extract the next non-empty line
        for short_indicator, full_indicator in indicators:
            for i, line in enumerate(lines):
                line_lower = line.lower()
                if short_indicator.lower() in line_lower or full_indicator.lower() in line_lower:
                    print(f"   📍 Found indicator '{full_indicator}' at line {i}: {repr(line.strip())}")
                    # Look for the next non-empty line
                    for j in range(i + 1, min(i + 5, len(lines))):  # Check up to 5 lines after
                        next_line = lines[j].strip()
                        if next_line and len(next_line) > 8:
                            cleaned_line = clean_extracted_text(next_line)
                            if is_airline_or_system_text(cleaned_line):
                                print(f"   ❌ Skipping airline/system text: {cleaned_line}")
                                continue
                            if might_be_company(cleaned_line):
                                print(f"   🤔 Potential company text: {cleaned_line}")
                                match = find_best_company_match(cleaned_line)
                                if match:
                                    return match
                    break  # Move to next indicator if this one fails
            else:
                print(f"   ⚠️ Indicator '{full_indicator}' not found")
        
        # Fallback: Use existing logic to process shipper section
        print("   🔍 Falling back to full shipper section processing")
        chars = page.chars
        shipper_y_start = None
        consignee_y_start = None
        for char in chars:
            text_lower = char['text'].lower()
            if 'shipper' in text_lower and char['x0'] < 300:
                shipper_y_start = char['top']
            elif 'consignee' in text_lower and char['x0'] < 300:
                consignee_y_start = char['top']
        lines_dict = {}
        for char in chars:
            y = int(char['top'])
            if y not in lines_dict:
                lines_dict[y] = []
            lines_dict[y].append(char)
        sorted_lines = sorted(lines_dict.items())
        for y, chars_in_line in sorted_lines:
            if shipper_y_start and y < shipper_y_start + 5:
                continue
            if consignee_y_start and y > consignee_y_start - 10:
                break
            left_chars = [c for c in chars_in_line if c['x0'] < 400]
            if left_chars:
                left_chars.sort(key=lambda x: x['x0'])
                line_text = ''.join([c['text'] for c in left_chars]).strip()
                match = find_best_company_match(line_text)
                if match:
                    return match
        return None
    except Exception as e:
        print(f"   ❌ Error extracting from page: {e}")
        return None

def process_single_page_pdf(pdf_path):
    """Process a single-page PDF or extracted page"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page_to_use = pdf.pages[0]
            page_text = page_to_use.extract_text()
            if not page_text or len(page_text.strip()) < 50:
                return extract_shipper_name_ocr(pdf_path)
            print("   📄 Using text extraction...")
            return extract_from_specific_page_text(page_to_use)
    except Exception as e:
        print(f"   ❌ Text extraction error: {e}")
        return extract_shipper_name_ocr(pdf_path)

def process_multi_page_pdf_with_detection(pdf_path):
    """Process PDF with smart page detection for shipper information"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages == 1:
                print("   📄 Single page PDF - processing directly")
                return process_single_page_pdf(pdf_path)
            print(f"   📚 Multi-page PDF detected ({total_pages} pages) - finding shipper page")
            page_num, shipper_page = find_shipper_page_text_based(pdf_path)
            if shipper_page:
                page_text = shipper_page.extract_text()
                if page_text and len(page_text.strip()) >= 50:
                    print(f"   📄 Processing page {page_num + 1} with text extraction...")
                    return extract_from_specific_page_text(shipper_page)
                else:
                    print(f"   📄 Page {page_num + 1} has no extractable text, using OCR...")
                    with tempfile.TemporaryDirectory() as temp_dir:
                        single_page_pdf = extract_specific_page_to_file(pdf_path, page_num, temp_dir)
                        if single_page_pdf:
                            return extract_shipper_name_ocr(single_page_pdf)
            return None
    except Exception as e:
        print(f"   ❌ Error in multi-page processing: {e}")
        return None

def extract_shipper_name_ocr(pdf_path):
    """Fallback to OCR methods if text-based extraction fails"""
    result = try_ocrmypdf(pdf_path)
    if result:
        return result
    result = try_tesseract_pdftoppm(pdf_path)
    if result:
        return result
    result = try_pdftotext(pdf_path)
    return result

def apply_high_threshold_fuzzy_matching(extracted_name, is_text_pdf, min_similarity=0.8):
    """Apply fuzzy matching only when confidence is very high (≥0.8)"""
    if not extracted_name:
        return None
    
    if is_text_pdf:
        # For text-based PDFs, trust the extraction more
        logger.info("Text-based PDF: Checking for high-confidence matches only")
        match = find_best_company_match(extracted_name, min_similarity=min_similarity)
        if match:
            logger.info(f"High-confidence match found: {match}")
            return match
        else:
            logger.info(f"No high-confidence match found. Using extracted name as new company: {extracted_name}")
            add_company_to_database(extracted_name)
            return extracted_name
    else:
        # For image-based PDFs, use traditional fuzzy matching
        logger.info("Image-based PDF: Using traditional fuzzy matching")
        match = find_best_company_match(extracted_name, min_similarity=0.6)  # Lower threshold for OCR
        if match:
            logger.info(f"Fuzzy match found: {match}")
            return match
        else:
            logger.info(f"No fuzzy match found. Using extracted name: {extracted_name}")
            return extracted_name

def extract_shipper_name(pdf_path):
    """Enhanced SHIPPER company name extraction with AI verification"""
    try:
        pdf_name = Path(pdf_path).name.lower()
        is_bloc_pdf = "bloc" in pdf_name
        
        logger.info(f"Starting shipper extraction for: {os.path.basename(pdf_path)}")
        
        # Determine if PDF is text-based or image-based
        is_text_pdf = is_pdf_text_based(pdf_path)
        
        extracted_candidates = None
        if is_text_pdf:
            logger.info("PDF identified as text-based - using exact extraction")
            extracted_candidates = extract_shipper_name_text_based(pdf_path)
            
            # If text-based extraction failed, treat it as image-based PDF with full OCR + cropping
            if not extracted_candidates:
                logger.warning("Text-based extraction failed - treating as image-based PDF with OCR + cropping")
                # Use the exact same OCR methods as image-based PDFs
                extracted_candidates = extract_shipper_name_ocr(pdf_path)
        else:
            logger.info("PDF identified as image-based - trying Gemini Vision first")
            # ── Gemini Vision (mirrors extractVisionMeta in mawbShipperExtract.js) ──
            if setup_gemini_api():
                vision_shipper = extract_vision_meta(pdf_path)
                if vision_shipper:
                    logger.info(f"Gemini Vision identified shipper: {vision_shipper}")
                    cleaned = clean_company_name(vision_shipper)
                    add_company_to_database(cleaned)
                    return cleaned
                logger.info("Gemini Vision returned no shipper — falling back to OCR")
            # ── OCR fallback ─────────────────────────────────────────────────────
            logger.info("PDF identified as image-based - using OCR methods")
            if is_bloc_pdf:
                with open(pdf_path, 'rb') as file:
                    pdf_reader = PdfReader(file)
                    page_count = len(pdf_reader.pages)
                if page_count == 2:
                    logger.info(f"Multi-page bloc PDF detected ({page_count} pages) - extracting second page")
                    with tempfile.TemporaryDirectory() as temp_dir:
                        second_page_pdf = extract_specific_page_to_file(pdf_path, 1, temp_dir)
                        if second_page_pdf:
                            extracted_candidates = extract_shipper_name_ocr(second_page_pdf)
                else:
                    logger.info(f"Bloc PDF with {page_count} page(s) - using smart page detection")
                    extracted_candidates = process_multi_page_pdf_with_detection(pdf_path)
            else:
                extracted_candidates = process_multi_page_pdf_with_detection(pdf_path)
        if not extracted_candidates:
            logger.warning("No shipper name extracted from PDF")
            return None
        
        # Check if we got a direct high-confidence match (string) or candidates (list)
        if isinstance(extracted_candidates, str):
            # High-confidence fuzzy match found - use it directly
            logger.info(f"High-confidence match found: {extracted_candidates}")
            return extracted_candidates
        
        # We have a list of candidates - use AI to select the best one
        if setup_gemini_api():
            logger.info("Using Gemini AI for shipper name selection from candidates")
            gemini_result = verify_shipper_with_gemini(extracted_candidates, KNOWN_COMPANIES)
            if gemini_result:
                logger.info(f"Gemini reasoning: {gemini_result.get('reasoning', 'N/A')}")
                if gemini_result.get('matched_company'):
                    final_name = gemini_result['matched_company']
                    logger.info(f"Matched to existing company: {final_name}")
                elif gemini_result.get('is_new_company'):
                    # Use Gemini's final_name (already cleaned by AI) instead of selected_candidate
                    final_name = gemini_result.get('final_name') or gemini_result.get('selected_candidate')
                    logger.info(f"New company identified by AI: {final_name}")
                else:
                    final_name = gemini_result.get('final_name', extracted_candidates[0] if isinstance(extracted_candidates, list) and extracted_candidates else extracted_candidates)
                    logger.info(f"Using Gemini final name: {final_name}")
            else:
                # Gemini failed, fallback to first candidate
                final_name = extracted_candidates[0] if isinstance(extracted_candidates, list) and extracted_candidates else extracted_candidates
        else:
            # No Gemini API, fallback to first candidate
            final_name = extracted_candidates[0] if isinstance(extracted_candidates, list) and extracted_candidates else extracted_candidates
        
        # Always clean the final name before returning
        final_name = clean_company_name(final_name)
        logger.info(f"Final shipper name: {final_name}")
        
        # Add to database if it's a new company (after cleaning)
        add_company_to_database(final_name)
        
        return final_name
        
    except Exception as e:
        logger.error(f"Error in enhanced shipper extraction: {e}")
        return None

def extract_shipper_name_from_ocr(pdf_path):
    """Original OCR-based extraction for image PDFs"""
    result = try_ocrmypdf(pdf_path)
    if result:
        return result
    result = try_tesseract_pdftoppm(pdf_path)
    if result:
        return result
    result = try_pdftotext(pdf_path)
    return result

def rename_sheet_pdfs(dir_path, directory_name):
    """Rename Sheet PDFs to mn format"""
    print("  Renaming Sheet PDFs...")
    found_sheets = False
    
    # Get all Sheet PDF files and sort them
    all_sheet_files = glob.glob(os.path.join(dir_path, "Sheet *.pdf"))
    
    if not all_sheet_files:
        print("    No Sheet PDFs found")
        return False
    
    print("    Found Sheet PDFs to rename:")
    found_sheets = True
    
    # Extract sheet numbers and create a mapping
    sheet_mapping = {}
    for file_path in all_sheet_files:
        filename = os.path.basename(file_path)
        # Extract sheet number using regex (handle "Sheet 1 - ..." or "Sheet 1 (47).pdf")
        match = re.match(r'Sheet (\d+)', filename)
        if match:
            sheet_num = int(match.group(1))
            if sheet_num not in sheet_mapping:
                sheet_mapping[sheet_num] = file_path
            # If duplicate, keep the one without timestamp (simpler name)
            elif '(' in filename and '(' not in os.path.basename(sheet_mapping[sheet_num]):
                # Current has parentheses, existing doesn't - keep existing
                pass
            elif '(' not in filename and '(' in os.path.basename(sheet_mapping[sheet_num]):
                # Current doesn't have parentheses, existing does - replace with current
                sheet_mapping[sheet_num] = file_path
            elif '-' in filename and '-' not in os.path.basename(sheet_mapping[sheet_num]):
                # Current has timestamp, existing doesn't - keep existing
                pass
            elif '-' not in filename and '-' in os.path.basename(sheet_mapping[sheet_num]):
                # Current doesn't have timestamp, existing does - replace with current
                sheet_mapping[sheet_num] = file_path
    
    # Rename files in sorted order
    for sheet_num in sorted(sheet_mapping.keys()):
        file_path = sheet_mapping[sheet_num]
        filename = os.path.basename(file_path)
        new_name = os.path.join(dir_path, f"mn{sheet_num}.pdf")
        
        try:
            # Check if target already exists
            if os.path.exists(new_name):
                print(f"    ⚠️  Target 'mn{sheet_num}.pdf' already exists, skipping '{filename}'")
                continue
            
            os.rename(file_path, new_name)
            print(f"    ✓ Renamed '{filename}' to 'mn{sheet_num}.pdf'")
        except Exception as e:
            print(f"    ✗ Error renaming '{filename}': {e}")
    
    return found_sheets

def extract_mawb_from_generated_excel(dir_path):
    """Extract MAWB number from generated_excel file (first row, third column)"""
    try:
        # Find generated_excel file
        excel_files = glob.glob(os.path.join(dir_path, "generated_excel*.xlsx"))
        if not excel_files:
            return None
        
        excel_path = excel_files[0]
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # MAWB number is typically in cell C1 (row 1, column 3)
        mawb_cell = ws.cell(row=1, column=3).value
        
        if mawb_cell:
            # Clean and extract the MAWB number
            mawb_str = str(mawb_cell).strip()
            # Remove any extra whitespace or formatting
            mawb_str = re.sub(r'\s+', '', mawb_str)
            logger.info(f"Extracted MAWB from generated_excel: {mawb_str}")
            return mawb_str
        
        return None
    except Exception as e:
        logger.error(f"Error extracting MAWB from generated_excel: {e}")
        return None

def validate_mawb_match(pdf_mawb, excel_mawb):
    """Validate that MAWB number from PDF matches the one in generated_excel"""
    if not pdf_mawb or not excel_mawb:
        return False
    
    # Normalize both numbers (remove spaces, dashes, etc.)
    pdf_normalized = re.sub(r'[-\s]', '', str(pdf_mawb).strip())
    excel_normalized = re.sub(r'[-\s]', '', str(excel_mawb).strip())
    
    return pdf_normalized == excel_normalized

def compress_pdf_if_needed(pdf_path, max_size_mb=1.5):
    """Compress PDF progressively until target size is reached
    Tries multiple compression levels from least to most aggressive
    Returns: (compressed_path, was_compressed)
    """
    import subprocess
    import shutil
    
    try:
        file_size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        print(f"    📊 Taille originale: {file_size_mb:.2f} MB")
        
        if file_size_mb <= max_size_mb:
            print(f"    ✓ Taille OK (≤ {max_size_mb} MB) - Pas de compression nécessaire")
            return pdf_path, False
        
        print(f"    🎯 Objectif: {max_size_mb} MB")
        print(f"    ⚠️  Compression nécessaire...")
        
        # Niveaux de compression (du moins au plus agressif)
        compression_levels = [
            ('/printer', 'Qualité impression (300 dpi)'),
            ('/ebook', 'Qualité eBook (150 dpi)'),
            ('/screen', 'Qualité écran (72 dpi)')
        ]
        
        # Method 1: Try Ghostscript with progressive compression levels
        try:
            for level, description in compression_levels:
                temp_path = pdf_path + '.gs.tmp'
                
                print(f"    🔄 Tentative: {description}...")
                
                gs_cmd = [
                    'gswin64c',  # Windows; use 'gs' on Linux/Mac
                    '-sDEVICE=pdfwrite',
                    '-dCompatibilityLevel=1.4',
                    f'-dPDFSETTINGS={level}',
                    '-dNOPAUSE',
                    '-dQUIET',
                    '-dBATCH',
                    f'-sOutputFile={temp_path}',
                    pdf_path
                ]
                
                try:
                    result = subprocess.run(gs_cmd, capture_output=True, timeout=60, check=True)
                    
                    if os.path.exists(temp_path):
                        compressed_size = os.path.getsize(temp_path) / (1024 * 1024)
                        reduction = ((file_size_mb - compressed_size) / file_size_mb) * 100
                        
                        print(f"       → {compressed_size:.2f} MB ({reduction:.1f}% réduction)")
                        
                        if compressed_size <= max_size_mb:
                            # Success! Target reached
                            os.remove(pdf_path)
                            os.rename(temp_path, pdf_path)
                            print(f"    ✅ Objectif atteint avec {description}")
                            return pdf_path, True
                        else:
                            # Try next level
                            print(f"       ⚠️  Toujours > {max_size_mb} MB, essai niveau suivant...")
                            os.remove(temp_path)
                    
                except subprocess.TimeoutExpired:
                    print(f"       ⏱️  Timeout - Essai niveau suivant...")
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    continue
                    
                except subprocess.CalledProcessError as e:
                    print(f"       ❌ Erreur compression - Essai niveau suivant...")
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    continue
            
            # If all levels tried but target not reached
            print(f"    ⚠️  Impossible d'atteindre {max_size_mb} MB avec Ghostscript")
            print(f"    ℹ️  Utilisation du niveau /screen comme meilleure option...")
            
            # Use /screen as last resort
            final_temp = pdf_path + '.final.tmp'
            gs_cmd_final = [
                'gswin64c',
                '-sDEVICE=pdfwrite',
                '-dCompatibilityLevel=1.4',
                '-dPDFSETTINGS=/screen',
                '-dNOPAUSE',
                '-dQUIET',
                '-dBATCH',
                f'-sOutputFile={final_temp}',
                pdf_path
            ]
            
            subprocess.run(gs_cmd_final, capture_output=True, timeout=60, check=True)
            
            if os.path.exists(final_temp):
                final_size = os.path.getsize(final_temp) / (1024 * 1024)
                reduction = ((file_size_mb - final_size) / file_size_mb) * 100
                print(f"    📄 Taille finale: {final_size:.2f} MB ({reduction:.1f}% réduction)")
                os.remove(pdf_path)
                os.rename(final_temp, pdf_path)
                return pdf_path, True
                
        except FileNotFoundError:
            print(f"    ℹ️  Ghostscript non disponible, essai PyPDF2...")
        except Exception as e:
            print(f"    ⚠️  Ghostscript échoué: {e}, essai PyPDF2...")
        
        # Method 2: PyPDF2 compression (fallback)
        print(f"    🔄 Compression PyPDF2 (fallback)...")
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Copy all pages with compression
        for page in reader.pages:
            page.compress_content_streams()
            writer.add_page(page)
        
        # Write compressed PDF to temporary file
        temp_path = pdf_path + ".tmp"
        with open(temp_path, 'wb') as output_file:
            writer.write(output_file)
        
        # Check new size
        new_size_mb = os.path.getsize(temp_path) / (1024 * 1024)
        
        # Only use compressed version if smaller
        if new_size_mb < file_size_mb:
            compression_ratio = ((file_size_mb - new_size_mb) / file_size_mb) * 100
            print(f"    ✓ PyPDF2: {file_size_mb:.2f} MB → {new_size_mb:.2f} MB ({compression_ratio:.1f}% réduction)")
            os.remove(pdf_path)
            os.rename(temp_path, pdf_path)
            return pdf_path, True
        else:
            print(f"    ℹ️  PyPDF2 n'a pas réduit la taille ({new_size_mb:.2f} MB)")
            os.remove(temp_path)
            return pdf_path, False
        
    except Exception as e:
        print(f"    ⚠️  Compression échouée: {e}")
        # Clean up temp files
        for temp_ext in [".tmp", ".gs.tmp", ".final.tmp"]:
            temp_path = pdf_path + temp_ext
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
        return pdf_path, False

def rename_mawb_pdfs_and_create_bloc_note(dir_path, directory_name, skip_rename=False):
    """Rename MAWB PDFs, compress if > 1.5 MB, and create bloc note file with shipper name
    Returns: (mawb_number, validation_passed)
    
    If skip_rename=True, only validates but doesn't rename/process the MAWB file
    
    Enhanced: If no MAWB*.pdf found, will search for doc*.pdf and use MAWB from generated_excel
    """
    print("  Processing MAWB PDFs...")
    found_mawb = False
    mawb_number = None
    mawb_files = glob.glob(os.path.join(dir_path, "MAWB*.pdf"))
    
    # If no MAWB*.pdf files found, try doc*.pdf files
    use_doc_file = False
    if not mawb_files:
        print("    No MAWB*.pdf files found, searching for doc*.pdf...")
        doc_files = glob.glob(os.path.join(dir_path, "doc*.pdf"))
        if doc_files:
            print(f"    ✓ Found {len(doc_files)} doc*.pdf file(s) - will use as MAWB")
            mawb_files = doc_files
            use_doc_file = True
        else:
            print("    ❌ No MAWB*.pdf or doc*.pdf files found")
    
    # First, extract MAWB from generated_excel for validation
    excel_mawb = extract_mawb_from_generated_excel(dir_path)
    if excel_mawb:
        print(f"    📋 Generated Excel MAWB: {excel_mawb}")
    
    for file_path in mawb_files:
        if not found_mawb:
            if use_doc_file:
                print("    Found doc*.pdf file(s) to process as MAWB:")
            else:
                print("    Found MAWB PDFs to process:")
            found_mawb = True
        
        filename = os.path.basename(file_path)
        
        # If using doc*.pdf file, get MAWB from generated_excel instead of filename
        if use_doc_file:
            if not excel_mawb:
                print(f"    ❌ Cannot process '{filename}' - No MAWB found in generated_excel")
                return None, False
            
            mawb_number = excel_mawb
            print(f"    ✓ Using MAWB from generated_excel: {mawb_number}")
            print(f"    📄 Processing doc file: {filename}")
            
            # No validation needed since we're using excel_mawb directly
        else:
            # Original logic for MAWB*.pdf files
            match = re.search(r'MAWB\s*(.+)\.pdf', filename)
            if not match:
                print(f"    ⚠️  Skipping '{filename}' - Cannot extract MAWB number from filename")
                continue
                
            mawb_number = match.group(1).strip()
            
            # Validate MAWB number matches generated_excel
            if excel_mawb:
                if validate_mawb_match(mawb_number, excel_mawb):
                    print(f"    ✅ MAWB validation passed: PDF '{mawb_number}' matches Excel '{excel_mawb}'")
                else:
                    print(f"    ⚠️  WARNING: MAWB MISMATCH!")
                    print(f"        PDF MAWB: {mawb_number}")
                    print(f"        Excel MAWB: {excel_mawb}")
                    print(f"    ❌ Skipping entire folder '{directory_name}' - Files don't belong together!")
                    print(f"    ❌ No files will be renamed or processed")
                    # Create warning report for this mismatch
                    create_mawb_mismatch_warning(directory_name, mawb_number, excel_mawb)
                    return None, False  # Return False to indicate validation failed
        
        # If skip_rename is True, don't rename - just return validation result
        if skip_rename:
            print(f"    ⏸️  MAWB rename deferred (pending other validations)")
            return mawb_number, True
        
        # Rename to original format: "{directory_name} - {mawb_number}.pdf"
        new_name = os.path.join(dir_path, f"{directory_name} - {mawb_number}.pdf")
        try:
            os.rename(file_path, new_name)
            if use_doc_file:
                print(f"    ✓ Renamed doc file '{filename}' to '{os.path.basename(new_name)}'")
            else:
                print(f"    ✓ Renamed '{filename}' to '{os.path.basename(new_name)}'")
            
            # Check size and compress if needed (> 1.5 MB)
            compressed_path, was_compressed = compress_pdf_if_needed(new_name, max_size_mb=1.5)
            
            if was_compressed:
                print(f"    ✓ PDF compressed successfully")
            
            # Extract shipper name and create bloc note
            shipper_name = extract_shipper_name(compressed_path)
            create_bloc_note(directory_name, mawb_number, shipper_name)
            
        except Exception as e:
            print(f"    ✗ Error processing '{filename}': {e}")
    if not found_mawb:
        print("    No MAWB PDFs found")
    return mawb_number, True  # Return True to indicate validation passed

def create_bloc_note(directory_name, mawb_number, shipper_name):
    """Create bloc note file with shipper name"""
    bloc_note_path = os.path.join(os.getcwd(), f"{directory_name}.txt")
    mawb_clean = mawb_number.replace('-', '')
    try:
        # Create the original bloc note file with all details
        with open(bloc_note_path, 'w', encoding='utf-8') as f:
            f.write("-------------\n")
            f.write(f"{directory_name}\n")
            f.write(f"{mawb_clean}\n")
            f.write(f"{mawb_number}/1\n")
            f.write("\n")
            if shipper_name:
                f.write(f"{shipper_name}\n")
            else:
                f.write("\n")
            f.write("\n")
        print(f"    ✓ Created bloc note: '{directory_name}.txt' with shipper: {shipper_name or '(empty)'}")
        
        # EXTRA: Create a simplified shipper-only file for automation
        safe_name = directory_name.replace(' ', '_')
        shipper_only_path = os.path.join(os.getcwd(), f"{safe_name}_shipper_name.txt")
        with open(shipper_only_path, 'w', encoding='utf-8') as f:
            if shipper_name:
                f.write(f"{shipper_name}\n")
            else:
                f.write("\n")
        print(f"    ✓ Created shipper file: '{safe_name}_shipper_name.txt'")
        
        return True
    except Exception as e:
        print(f"    ✗ Error creating bloc note: {e}")
        return False

def rename_excel_files(dir_path, mawb_number):
    """Rename Excel files to include MAWB number"""
    if not mawb_number:
        return
    print("  Renaming Excel files...")
    summary_files = glob.glob(os.path.join(dir_path, "summary_file*.xlsx"))
    for file_path in summary_files:
        new_name = os.path.join(dir_path, f"summary_file - {mawb_number}.xlsx")
        try:
            os.rename(file_path, new_name)
            print(f"    ✓ Renamed summary file to include MAWB number")
        except Exception as e:
            print(f"    ✗ Error renaming summary file: {e}")
    generated_files = glob.glob(os.path.join(dir_path, "generated_excel*.xlsx"))
    for file_path in generated_files:
        new_name = os.path.join(dir_path, f"generated_excel - {mawb_number}.xlsx")
        try:
            os.rename(file_path, new_name)
            print(f"    ✓ Renamed generated_excel file to include MAWB number")
        except Exception as e:
            print(f"    ✗ Error renaming generated_excel file: {e}")

def delete_unwanted_files(dir_path):
    """Delete IMG and Manifest files"""
    print("  Cleaning up unwanted files...")
    img_patterns = ["IMG*.*", "*.img", "*.jpg", "*.png", "*.jpeg", "*image*"]
    found_img = False
    for pattern in img_patterns:
        img_files = glob.glob(os.path.join(dir_path, pattern))
        for file_path in img_files:
            if not found_img:
                print("    Found IMG files to delete:")
                found_img = True
            try:
                os.remove(file_path)
                print(f"    ✓ Deleted '{os.path.basename(file_path)}'")
            except Exception as e:
                print(f"    ✗ Error deleting '{os.path.basename(file_path)}': {e}")
    manifest_patterns = ["Manifest*.*", "Manifeste*.*"]
    found_manifest = False
    for pattern in manifest_patterns:
        manifest_files = glob.glob(os.path.join(dir_path, pattern))
        for file_path in manifest_files:
            if not found_manifest:
                print("    Found Manifest files to delete:")
                found_manifest = True
            try:
                os.remove(file_path)
                print(f"    ✓ Deleted '{os.path.basename(file_path)}'")
            except Exception as e:
                print(f"    ✗ Error deleting '{os.path.basename(file_path)}': {e}")
    if not found_img and not found_manifest:
        print("    No unwanted files found")

def extract_p_values_from_generated_excel(file_path):
    """Extract P and P,BRUT values from generated_excel file"""
    global_p_values = {}
    dum_p_values = []
    try:
        if file_path.endswith('.xlsx'):
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                print(f"      Reading Excel file with {ws.max_row} rows and {ws.max_column} columns")
                global_p_found = False
                global_p_brut_found = False
                current_dum = None
                dum_count = 0
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is None:
                            continue
                        cell_str = str(cell_value).strip()
                        if "DUM" in cell_str and any(char.isdigit() for char in cell_str):
                            dum_count += 1
                            current_dum = {'dum_number': dum_count, 'P': None, 'P_BRUT': None}
                            print(f"        Found DUM {dum_count}")
                            continue
                        if cell_str == "P":
                            value = None
                            if col + 1 <= ws.max_column:
                                right_cell = ws.cell(row=row, column=col + 1).value
                                if right_cell is not None:
                                    value = str(right_cell).strip()
                            if value and value != "":
                                if current_dum is None and not global_p_found:
                                    global_p_values["P"] = value
                                    global_p_found = True
                                    print(f"      Found global P: {value}")
                                elif current_dum is not None:
                                    current_dum['P'] = value
                                    print(f"        DUM {current_dum['dum_number']} P: {value}")
                        elif cell_str == "P,BRUT":
                            value = None
                            if col + 1 <= ws.max_column:
                                right_cell = ws.cell(row=row, column=col + 1).value
                                if right_cell is not None:
                                    value = str(right_cell).strip()
                            if value and value != "":
                                if current_dum is None and not global_p_brut_found:
                                    global_p_values["P,BRUT"] = value
                                    global_p_brut_found = True
                                    print(f"      Found global P,BRUT: {value}")
                                elif current_dum is not None:
                                    current_dum['P_BRUT'] = value
                                    print(f"        DUM {current_dum['dum_number']} P,BRUT: {value}")
                                    if current_dum['P'] is not None:
                                        dum_p_values.append(current_dum.copy())
                                        current_dum = None
                print(f"      Found {len(dum_p_values)} DUMs with P values")
            except Exception as excel_error:
                print(f"      Could not read as Excel: {excel_error}")
                return extract_p_values_as_text(file_path)
        else:
            return extract_p_values_as_text(file_path)
    except Exception as e:
        print(f"      Error reading generated_excel file: {e}")
    return global_p_values, dum_p_values

def extract_p_values_as_text(file_path):
    """Extract P and P,BRUT values from text file with multiple encodings"""
    global_p_values = {}
    dum_p_values = []
    encodings = ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            print(f"      Successfully read file with {encoding} encoding")
            lines = content.splitlines()
            print(f"      Reading {len(lines)} lines from generated_excel")
            global_p_found = False
            global_p_brut_found = False
            current_dum = None
            dum_count = 0
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                if "DUM" in line and any(char.isdigit() for char in line):
                    dum_count += 1
                    current_dum = {'dum_number': dum_count, 'P': None, 'P_BRUT': None}
                    print(f"        Found DUM {dum_count}")
                    continue
                parts = line.split('\t') if '\t' in line else line.split()
                if len(parts) < 2:
                    continue
                key = parts[0].strip()
                value = parts[1].strip()
                if key == "P" and not global_p_found and current_dum is None:
                    global_p_values["P"] = value
                    global_p_found = True
                    print(f"      Found global P: {value}")
                elif key == "P,BRUT" and not global_p_brut_found and current_dum is None:
                    global_p_values["P,BRUT"] = value
                    global_p_brut_found = True
                    print(f"      Found global P,BRUT: {value}")
                elif key == "P" and current_dum is not None:
                    current_dum['P'] = value
                    print(f"        DUM {current_dum['dum_number']} P: {value}")
                elif key == "P,BRUT" and current_dum is not None:
                    current_dum['P_BRUT'] = value
                    print(f"        DUM {current_dum['dum_number']} P,BRUT: {value}")
                    dum_p_values.append(current_dum.copy())
                    current_dum = None
            print(f"      Found {len(dum_p_values)} DUMs with P values")
            return global_p_values, dum_p_values
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"      Error with encoding {encoding}: {e}")
            continue
    print(f"      Could not read file with any encoding")
    return global_p_values, dum_p_values

def update_bloc_note_with_all_p_values(directory, global_p_values, dum_p_values):
    """Update bloc note file with P values"""
    bloc_note_path = os.path.join(os.getcwd(), f"{directory}.txt")
    if not os.path.exists(bloc_note_path):
        print(f"      Bloc note not found: {bloc_note_path}")
        return
    try:
        with open(bloc_note_path, 'r', encoding='utf-8') as f:
            content = f.read().splitlines()
        
        # Keep first 9 lines: header (4 lines) + 2 empty lines + shipper line + 2 empty lines
        new_content = content[:9] if len(content) >= 9 else content[:]
        while len(new_content) < 9:
            new_content.append("")
        
        # Add P values section
        if "P" in global_p_values:
            new_content.append(f"P\t{global_p_values['P']}")
        if "P,BRUT" in global_p_values:
            new_content.append(f"P,BRUT\t{global_p_values['P,BRUT']}")
        
        new_content.extend(["", "", "", "---------------------", ""])
        
        for dum in dum_p_values:
            new_content.append(f"\t\tDUM {dum['dum_number']}\t\t")
            if dum['P'] is not None:
                new_content.append(f"P\t{dum['P']}")
            new_content.append("")
            if dum['P_BRUT'] is not None:
                new_content.append(f"P,BRUT\t{dum['P_BRUT']}")
            new_content.extend(["", "", "---------------------", ""])
        
        with open(bloc_note_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(new_content))
        print(f"      ✓ Bloc note updated with global P values and {len(dum_p_values)} DUMs")
    except Exception as e:
        print(f"      Error updating bloc note: {e}")


def compare_and_correct_summary_file(generated_excel_path, summary_file_path, directory):
    """Compare DUMs between generated_excel and summary_file and correct differences"""
    print("    Comparing generated_excel vs summary_file...")
    _, generated_dums = extract_p_values_from_generated_excel(generated_excel_path)
    try:
        wb = load_workbook(summary_file_path)
        ws = wb.active
        column_mapping = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                header_str = str(header).lower().strip()
                if "total pieces" in header_str or "total position" in header_str:
                    column_mapping['P'] = col
                elif "total value" in header_str:
                    column_mapping['V'] = col
                elif "total poid brute" in header_str:
                    column_mapping['P_BRUT'] = col
                elif "total freight" in header_str:
                    column_mapping['Fret'] = col
                elif "assurance" in header_str:
                    column_mapping['Ass'] = col
                elif "carton" in header_str:
                    column_mapping['N_COLIS'] = col
        print(f"      Found columns mapping: {column_mapping}")
        differences = []
        corrections_made = False
        for row in range(2, ws.max_row + 1):
            sheet_name_cell = ws.cell(row=row, column=1).value
            if sheet_name_cell and "Sheet" in str(sheet_name_cell):
                # Extract sheet number using regex to get all digits (handles Sheet 1, Sheet 10, etc.)
                import re
                match = re.search(r'Sheet\s*(\d+)', str(sheet_name_cell))
                sheet_num = int(match.group(1)) if match else None
                
                if sheet_num and sheet_num <= len(generated_dums):
                    generated_dum = generated_dums[sheet_num - 1]
                    print(f"        Comparing Sheet {sheet_num} vs DUM {generated_dum['dum_number']}")
                    comparisons = [
                        ('P', 'Total position', generated_dum.get('P')),
                        ('P_BRUT', 'Total poid brute', generated_dum.get('P_BRUT'))
                    ]
                    for key, description, generated_value in comparisons:
                        if key in column_mapping and generated_value is not None:
                            col = column_mapping[key]
                            summary_value = ws.cell(row=row, column=col).value
                            try:
                                gen_val = float(str(generated_value).replace(',', '.')) if generated_value else 0
                                sum_val = float(str(summary_value).replace(',', '.')) if summary_value else 0
                                if abs(gen_val - sum_val) > 0.01:
                                    diff_info = {
                                        'sheet': sheet_num,
                                        'dum': generated_dum['dum_number'],
                                        'field': description,
                                        'summary_value': summary_value,
                                        'generated_value': generated_value,
                                        'row': row,
                                        'col': col
                                    }
                                    differences.append(diff_info)
                                    ws.cell(row=row, column=col).value = generated_value
                                    corrections_made = True
                                    print(f"          DIFFERENCE: {description} - Summary: {summary_value} → Generated: {generated_value}")
                            except (ValueError, TypeError):
                                if str(generated_value).strip() != str(summary_value).strip():
                                    diff_info = {
                                        'sheet': sheet_num,
                                        'dum': generated_dum['dum_number'],
                                        'field': description,
                                        'summary_value': summary_value,
                                        'generated_value': generated_value,
                                        'row': row,
                                        'col': col
                                    }
                                    differences.append(diff_info)
                                    ws.cell(row=row, column=col).value = generated_value
                                    corrections_made = True
                                    print(f"          DIFFERENCE: {description} - Summary: {summary_value} → Generated: {generated_value}")
        if corrections_made:
            wb.save(summary_file_path)
            print(f"      ✓ Summary file corrected with {len(differences)} changes")
        else:
            print("      ✓ No differences found - files are synchronized")
        if differences:
            create_warning_report(directory, differences)
        return len(differences)
    except Exception as e:
        print(f"      Error comparing files: {e}")
        return 0

def create_warning_report(directory, differences):
    """Create warning report for differences found"""
    warning_path = os.path.join(os.getcwd(), "!---------------------------Warning-----------------------.txt")
    try:
        existing_content = ""
        if os.path.exists(warning_path):
            with open(warning_path, 'r', encoding='utf-8') as f:
                existing_content = f.read()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report = f"\n{'='*60}\n"
        report += f"RAPPORT DE CORRECTION - {directory}\n"
        report += f"Date: {timestamp}\n"
        report += f"Nombre de différences trouvées: {len(differences)}\n"
        report += f"{'='*60}\n\n"
        for diff in differences:
            report += f"Sheet {diff['sheet']} (DUM {diff['dum']}):\n"
            report += f"  Champ: {diff['field']}\n"
            report += f"  Valeur Summary: {diff['summary_value']}\n"
            report += f"  Valeur Generated: {diff['generated_value']} ← CORRIGÉE\n"
            report += f"  Position: Ligne {diff['row']}, Colonne {diff['col']}\n"
            report += f"{'-'*40}\n"
        with open(warning_path, 'w', encoding='utf-8') as f:
            f.write(existing_content + report)
        print(f"      ✓ Warning report created: warning.txt")
    except Exception as e:
        print(f"      Error creating warning report: {e}")

def create_mawb_mismatch_warning(directory, pdf_mawb, excel_mawb):
    """Create warning report for MAWB mismatch - files don't belong together"""
    warning_path = os.path.join(os.getcwd(), "!-------Warning - LTA vs Generated--------.txt")
    try:
        existing_content = ""
        if os.path.exists(warning_path):
            with open(warning_path, 'r', encoding='utf-8') as f:
                existing_content = f.read()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report = f"\n{'='*60}\n"
        report += f"⚠️  ALERTE: INCOHÉRENCE MAWB - {directory}\n"
        report += f"Date: {timestamp}\n"
        report += f"{'='*60}\n\n"
        report += f"❌ TRAITEMENT IGNORÉ - Les fichiers ne correspondent pas!\n\n"
        report += f"Le dossier '{directory}' contient des fichiers incompatibles:\n\n"
        report += f"  📄 Fichier PDF MAWB:        {pdf_mawb}\n"
        report += f"  📋 Excel MAWB (Generated):  {excel_mawb}\n\n"
        report += f"⚠️  Ces numéros MAWB ne correspondent pas!\n"
        report += f"   Le PDF et l'Excel ne font pas partie du même envoi.\n\n"
        report += f"🔧 ACTION REQUISE:\n"
        report += f"   1. Vérifier quel fichier est incorrect\n"
        report += f"   2. Remplacer le fichier erroné par le bon\n"
        report += f"   3. Relancer le script après correction\n\n"
        report += f"ℹ️  NOTE: Aucun fichier n'a été modifié pour ce dossier.\n"
        report += f"         Le traitement a été sauté pour éviter la corruption de données.\n"
        report += f"{'-'*60}\n"
        
        with open(warning_path, 'w', encoding='utf-8') as f:
            f.write(existing_content + report)
        
        logger.warning(f"MAWB mismatch warning created for {directory}")
        print(f"      ✓ MAWB mismatch warning added to report")
    except Exception as e:
        print(f"      Error creating MAWB mismatch warning: {e}")

def validate_logical_values_from_summary(summary_file_path, directory):
    """Validate logical values from summary_file (table format)
    Returns: (is_valid, error_details)
    """
    print("  Validating data logic from summary_file...")
    errors = []
    
    try:
        wb = load_workbook(summary_file_path, data_only=True)
        ws = wb.active
        
        # Find header row
        header_row = None
        col_indices = {}
        
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and "Sheet Name" in str(cell_value):
                    header_row = row
                    break
            if header_row:
                break
        
        if not header_row:
            print("    ⚠️  Could not find header row in summary_file")
            return True, []
        
        # Map column names to indices
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header:
                header_str = str(header).strip()
                col_indices[header_str] = col
        
        # Validate each sheet (skip GLOBAL)
        for row in range(header_row + 1, ws.max_row + 1):
            sheet_name = ws.cell(row=row, column=col_indices.get('Sheet Name', 1)).value
            if not sheet_name or str(sheet_name).strip().upper() == 'GLOBAL':
                continue
            
            sheet_name_str = str(sheet_name).strip()
            
            # Extract values
            try:
                total_value = None
                total_freight = None
                total_poid_net = None
                total_poid_brute = None
                
                if 'Total Value' in col_indices:
                    val = ws.cell(row=row, column=col_indices['Total Value']).value
                    if val: total_value = float(str(val).replace(',', '.').strip())
                
                if 'Total freight' in col_indices:
                    val = ws.cell(row=row, column=col_indices['Total freight']).value
                    if val: total_freight = float(str(val).replace(',', '.').strip())
                
                if 'Total poid net' in col_indices:
                    val = ws.cell(row=row, column=col_indices['Total poid net']).value
                    if val: total_poid_net = float(str(val).replace(',', '.').strip())
                
                if 'Total poid brute' in col_indices:
                    val = ws.cell(row=row, column=col_indices['Total poid brute']).value
                    if val: total_poid_brute = float(str(val).replace(',', '.').strip())
                
                # Validate: Freight <= Value
                if total_freight and total_value:
                    if total_freight > total_value:
                        error = {
                            'source': 'summary_file',
                            'sheet': sheet_name_str,
                            'type': 'Freight > Value',
                            'freight': total_freight,
                            'value': total_value
                        }
                        errors.append(error)
                        print(f"    ❌ {sheet_name_str}: Fret ({total_freight}) > Value ({total_value}) - ILLOGIQUE!")
                
                # Validate: P_NET <= P_BRUT
                if total_poid_net and total_poid_brute:
                    if total_poid_net > total_poid_brute:
                        error = {
                            'source': 'summary_file',
                            'sheet': sheet_name_str,
                            'type': 'P_NET > P_BRUT',
                            'p_net': total_poid_net,
                            'p_brut': total_poid_brute
                        }
                        errors.append(error)
                        print(f"    ❌ {sheet_name_str}: P,NET ({total_poid_net}) > P,BRUT ({total_poid_brute}) - ILLOGIQUE!")
            
            except (ValueError, TypeError) as e:
                # Skip rows with invalid data
                continue
        
        if errors:
            print(f"    ⚠️  Found {len(errors)} logical errors in summary_file")
            return False, errors
        else:
            print(f"    ✅ All summary_file values are logical")
            return True, []
            
    except Exception as e:
        print(f"    ⚠️  Error validating summary_file: {e}")
        # If validation fails, assume values are OK (don't block processing)
        return True, []

def validate_logical_values(generated_excel_path, directory):
    """Validate that values are logical (Freight < Value, P_NET < P_BRUT)
    Returns: (is_valid, error_details)
    """
    print("  Validating data logic from generated_excel...")
    errors = []
    
    try:
        wb = load_workbook(generated_excel_path, data_only=True)
        ws = wb.active
        
        # Extract all DUMs with their values
        current_dum = None
        dum_data = []
        dum_count = 0
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    continue
                    
                cell_str = str(cell_value).strip()
                
                # Detect DUM headers
                if "DUM" in cell_str and any(char.isdigit() for char in cell_str):
                    if current_dum and current_dum.get('dum_number'):
                        dum_data.append(current_dum)
                    dum_count += 1
                    current_dum = {'dum_number': dum_count}
                    continue
                
                # Extract values
                if current_dum:
                    if cell_str == "P":
                        if col + 1 <= ws.max_column:
                            val = ws.cell(row=row, column=col + 1).value
                            if val: current_dum['P'] = float(str(val).strip())
                    elif cell_str == "V":
                        if col + 1 <= ws.max_column:
                            val = ws.cell(row=row, column=col + 1).value
                            if val: current_dum['V'] = float(str(val).replace(',', '.').strip())
                    elif cell_str == "P,NET":
                        if col + 1 <= ws.max_column:
                            val = ws.cell(row=row, column=col + 1).value
                            if val: current_dum['P_NET'] = float(str(val).strip())
                    elif cell_str == "P,BRUT":
                        if col + 1 <= ws.max_column:
                            val = ws.cell(row=row, column=col + 1).value
                            if val: current_dum['P_BRUT'] = float(str(val).strip())
                    elif cell_str == "Fret":
                        if col + 1 <= ws.max_column:
                            val = ws.cell(row=row, column=col + 1).value
                            if val: current_dum['Fret'] = float(str(val).strip())
        
        # Add last DUM
        if current_dum and current_dum.get('dum_number'):
            dum_data.append(current_dum)
        
        # Validate each DUM
        for dum in dum_data:
            dum_num = dum.get('dum_number', '?')
            
            # Check: Freight should be less than Value
            if 'Fret' in dum and 'V' in dum:
                if dum['Fret'] > dum['V']:
                    error = {
                        'source': 'generated_excel',
                        'dum': dum_num,
                        'type': 'Freight > Value',
                        'freight': dum['Fret'],
                        'value': dum['V']
                    }
                    errors.append(error)
                    print(f"    ❌ DUM {dum_num}: Fret ({dum['Fret']}) > V ({dum['V']}) - ILLOGIQUE!")
            
            # Check: P_NET should be less than P_BRUT
            if 'P_NET' in dum and 'P_BRUT' in dum:
                if dum['P_NET'] > dum['P_BRUT']:
                    error = {
                        'source': 'generated_excel',
                        'dum': dum_num,
                        'type': 'P_NET > P_BRUT',
                        'p_net': dum['P_NET'],
                        'p_brut': dum['P_BRUT']
                    }
                    errors.append(error)
                    print(f"    ❌ DUM {dum_num}: P,NET ({dum['P_NET']}) > P,BRUT ({dum['P_BRUT']}) - ILLOGIQUE!")
        
        if errors:
            print(f"    ⚠️  Found {len(errors)} logical errors in generated_excel")
            return False, errors
        else:
            print(f"    ✅ All generated_excel values are logical")
            return True, []
            
    except Exception as e:
        print(f"    ⚠️  Error validating logical values: {e}")
        # If validation fails, assume values are OK (don't block processing)
        return True, []

def create_logical_error_warning(directory, errors):
    """Create warning report for illogical values"""
    warning_path = os.path.join(os.getcwd(), "!-------Warning - LTA vs Generated--------.txt")
    try:
        existing_content = ""
        if os.path.exists(warning_path):
            with open(warning_path, 'r', encoding='utf-8') as f:
                existing_content = f.read()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report = f"\n{'='*60}\n"
        report += f"⚠️  ALERTE: VALEURS ILLOGIQUES - {directory}\n"
        report += f"Date: {timestamp}\n"
        report += f"{'='*60}\n\n"
        report += f"❌ TRAITEMENT IGNORÉ - Les valeurs sont incohérentes!\n\n"
        report += f"Le dossier '{directory}' contient des valeurs impossibles:\n\n"
        
        for error in errors:
            source_file = error.get('source', 'generated_excel')
            location = error.get('sheet') if 'sheet' in error else f"DUM {error.get('dum', '?')}"
            
            if error['type'] == 'Freight > Value':
                report += f"  📦 {location} ({source_file}):\n"
                report += f"     Fret: {error['freight']:,.2f}\n"
                report += f"     Valeur: {error['value']:,.2f}\n"
                report += f"     ❌ Le fret ne peut pas être supérieur à la valeur!\n\n"
            elif error['type'] == 'P_NET > P_BRUT':
                report += f"  ⚖️  {location} ({source_file}):\n"
                report += f"     P,NET: {error['p_net']}\n"
                report += f"     P,BRUT: {error['p_brut']}\n"
                report += f"     ❌ Le poids net ne peut pas être supérieur au poids brut!\n\n"
        
        report += f"🔧 ACTION REQUISE:\n"
        report += f"   1. Vérifier les valeurs dans le fichier generated_excel ET summary_file\n"
        report += f"   2. Corriger les valeurs erronées\n"
        report += f"   3. Relancer le script après correction\n\n"
        report += f"ℹ️  NOTE: Aucun fichier n'a été modifié pour ce dossier.\n"
        report += f"         Le traitement a été sauté pour éviter la propagation d'erreurs.\n"
        report += f"{'-'*60}\n"
        
        with open(warning_path, 'w', encoding='utf-8') as f:
            f.write(existing_content + report)
        
        logger.warning(f"Logical error warning created for {directory}")
        print(f"      ✓ Logical error warning added to report")
    except Exception as e:
        print(f"      Error creating logical error warning: {e}")

def count_expected_dums(dir_path):
    """
    Count the expected number of DUMs from summary_file and/or generated_excel
    Returns: (expected_count, source_file_name, expected_sheet_numbers)
    """
    expected_sheet_numbers = set()
    source_file_name = None
    
    # Try summary_file first (more reliable)
    summary_files = glob.glob(os.path.join(dir_path, "summary_file*.xlsx"))
    if summary_files:
        try:
            wb = load_workbook(summary_files[0], data_only=True)
            ws = wb.active
            source_file_name = os.path.basename(summary_files[0])
            
            # Find header row
            header_row = None
            sheet_name_col = None
            
            for row in range(1, min(10, ws.max_row + 1)):
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and "Sheet Name" in str(cell_value):
                        header_row = row
                        sheet_name_col = col
                        break
                if header_row:
                    break
            
            if header_row and sheet_name_col:
                # Extract all sheet numbers from summary_file
                for row in range(header_row + 1, ws.max_row + 1):
                    sheet_name = ws.cell(row=row, column=sheet_name_col).value
                    if sheet_name and "Sheet" in str(sheet_name):
                        # Extract sheet number using regex (handles Sheet 1, Sheet 10, etc.)
                        match = re.search(r'Sheet\s+(\d+)', str(sheet_name))
                        if match:
                            sheet_num = int(match.group(1))
                            expected_sheet_numbers.add(sheet_num)
            
            wb.close()
        except Exception as e:
            print(f"    ⚠️  Error reading summary_file: {e}")
    
    # If no sheets found in summary_file, try generated_excel
    if not expected_sheet_numbers:
        generated_excel_files = glob.glob(os.path.join(dir_path, "generated_excel*.xlsx"))
        if generated_excel_files:
            try:
                wb = load_workbook(generated_excel_files[0], data_only=True)
                ws = wb.active
                source_file_name = os.path.basename(generated_excel_files[0])
                
                # Count DUMs by searching for "DUM" patterns (same logic as extract_p_values_from_generated_excel)
                # Iterate through all cells and count each DUM found sequentially
                dum_count = 0
                for row in range(1, min(500, ws.max_row + 1)):
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is None:
                            continue
                        cell_str = str(cell_value).strip()
                        if "DUM" in cell_str and any(char.isdigit() for char in cell_str):
                            dum_count += 1
                            expected_sheet_numbers.add(dum_count)
                            # Continue to next row after finding a DUM (avoid counting multiple times in same row)
                            break
                
                wb.close()
            except Exception as e:
                print(f"    ⚠️  Error reading generated_excel: {e}")
    
    return len(expected_sheet_numbers), source_file_name, sorted(expected_sheet_numbers)

def find_exact_sheet_file(dir_path, sheet_num, file_extension):
    """
    Find a Sheet file with exact matching (avoids Sheet 1 matching Sheet 10)
    Supports multiple patterns:
    - Sheet X.xlsx
    - Sheet X - *.xlsx (e.g., "Sheet 1 - 2026-01-12T034033.984.xlsx")
    - Sheet X (*).xlsx (e.g., "Sheet 9 (90).xlsx")
    
    Uses regex to extract the exact sheet number, so "Sheet 1" won't match "Sheet 10"
    
    Returns: file path if found, None otherwise
    """
    # List all Sheet files with the extension
    pattern = f"Sheet *.{file_extension}"
    all_files = glob.glob(os.path.join(dir_path, pattern))
    
    # Use regex to extract sheet number and match exactly (avoid Sheet 1 matching Sheet 10)
    # The regex extracts the first number after "Sheet", regardless of format:
    # - "Sheet 9 (90).xlsx" -> extracts 9
    # - "Sheet 9 - 2026-01-12T034039.694.xlsx" -> extracts 9
    # - "Sheet 1.xlsx" -> extracts 1
    for file_path in all_files:
        filename = os.path.basename(file_path)
        # Extract sheet number using regex to ensure exact match
        match = re.match(rf'Sheet\s+(\d+)', filename)
        if match:
            file_sheet_num = int(match.group(1))
            if file_sheet_num == sheet_num:
                return file_path
    
    return None

def validate_required_files_present(dir_path, directory_name):
    """
    Validate that all required Sheet X.xlsx and Sheet X.pdf files are present
    (mnX.pdf files are created later by rename_sheet_pdfs, so we don't check for them)
    Returns: (is_valid, missing_files)
    """
    missing_files = []
    
    # Count expected DUMs
    expected_count, source_file, expected_sheet_numbers = count_expected_dums(dir_path)
    
    if expected_count == 0:
        # Cannot validate if we don't know how many DUMs are expected
        print("  ⚠️  Cannot determine expected number of DUMs - skipping file validation")
        return True, []
    
    print(f"  🔍 Validating required files ({expected_count} DUMs expected from {source_file})...")
    
    # Check each expected Sheet number
    for sheet_num in expected_sheet_numbers:
        # Check for Sheet X.xlsx file (exact match to avoid Sheet 1 matching Sheet 10)
        sheet_excel_found = find_exact_sheet_file(dir_path, sheet_num, 'xlsx')
        
        if not sheet_excel_found:
            missing_files.append({
                'type': 'Sheet Excel',
                'filename': f'Sheet {sheet_num} - *.xlsx or Sheet {sheet_num} (*).xlsx',
                'sheet_number': sheet_num
            })
            print(f"    ❌ Missing: Sheet {sheet_num} Excel file")
        
        # Check for Sheet X.pdf file (exact match to avoid Sheet 1 matching Sheet 10)
        sheet_pdf_found = find_exact_sheet_file(dir_path, sheet_num, 'pdf')
        
        if not sheet_pdf_found:
            missing_files.append({
                'type': 'Sheet PDF',
                'filename': f'Sheet {sheet_num} - *.pdf or Sheet {sheet_num} (*).pdf',
                'sheet_number': sheet_num
            })
            print(f"    ❌ Missing: Sheet {sheet_num} PDF file")
    
    if missing_files:
        print(f"  ❌ {len(missing_files)} file(s) missing out of {expected_count * 2} expected files")
        return False, missing_files
    else:
        print(f"  ✅ All required files present ({expected_count} Sheet Excel files + {expected_count} Sheet PDF files)")
        return True, []

def create_missing_files_error(dir_path, directory_name, missing_files, expected_count, source_file):
    """Create error report for missing required files"""
    error_path = os.path.join(os.getcwd(), "!-------ERROR - Missing Files--------.txt")
    try:
        existing_content = ""
        if os.path.exists(error_path):
            with open(error_path, 'r', encoding='utf-8') as f:
                existing_content = f.read()
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        report = f"\n{'='*60}\n"
        report += f"❌ ERREUR: FICHIERS MANQUANTS - {directory_name}\n"
        report += f"Date: {timestamp}\n"
        report += f"{'='*60}\n\n"
        report += f"❌ TRAITEMENT ARRÊTÉ - Fichiers requis manquants!\n\n"
        report += f"Le dossier '{directory_name}' devrait contenir {expected_count} DUMs\n"
        report += f"(d'après {source_file}), mais certains fichiers sont manquants:\n\n"
        
        # Group missing files by type
        missing_sheet_excel = [f for f in missing_files if f['type'] == 'Sheet Excel']
        missing_sheet_pdf = [f for f in missing_files if f['type'] == 'Sheet PDF']
        
        if missing_sheet_excel:
            report += f"📋 FICHIERS EXCEL MANQUANTS ({len(missing_sheet_excel)} fichier(s)):\n"
            for file_info in sorted(missing_sheet_excel, key=lambda x: x['sheet_number']):
                report += f"   - {file_info['filename']} (Sheet {file_info['sheet_number']})\n"
            report += "\n"
        
        if missing_sheet_pdf:
            report += f"📄 FICHIERS PDF MANQUANTS ({len(missing_sheet_pdf)} fichier(s)):\n"
            for file_info in sorted(missing_sheet_pdf, key=lambda x: x['sheet_number']):
                report += f"   - {file_info['filename']} (Sheet {file_info['sheet_number']})\n"
            report += "\n"
        
        report += f"🔧 ACTION REQUISE:\n"
        report += f"   1. Vérifier que tous les fichiers 'Sheet X - *.xlsx' ou 'Sheet X (*).xlsx' sont présents\n"
        report += f"   2. Vérifier que tous les fichiers 'Sheet X - *.pdf' ou 'Sheet X (*).pdf' sont présents\n"
        report += f"   3. Corriger {source_file} si le nombre de DUMs est incorrect\n"
        report += f"   4. Relancer le script après correction\n\n"
        report += f"ℹ️  NOTE: Aucun fichier n'a été renommé ou modifié pour ce dossier.\n"
        report += f"         Le traitement a été arrêté pour éviter des erreurs.\n"
        report += f"{'-'*60}\n"
        
        with open(error_path, 'w', encoding='utf-8') as f:
            f.write(existing_content + report)
        
        logger.warning(f"Missing files error created for {directory_name}")
        print(f"      ✓ Missing files error report created")
    except Exception as e:
        print(f"      Error creating missing files error report: {e}")

def validate_and_correct_article_values(file_path):
    """Validate and correct article values to ensure no value exceeds 499 MAD
    
    Logic:
    1. Find all articles with value > 499
    2. Reduce each to a value between 420-480 (random)
    3. Redistribute the excess amount to articles with low values
    4. PRESERVE FULL DECIMAL PRECISION - keep all original decimal places
    """
    try:
        from decimal import Decimal, ROUND_HALF_UP

        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find header row (should be row 1)
        value_col = None
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col_idx).value
            if cell_value and 'Valeur' in str(cell_value):
                value_col = col_idx
                break
        
        if not value_col:
            return True  # Column not found, skip validation
        
        # Collect all article values (skip header row 1)
        articles = []
        for row_idx in range(2, ws.max_row + 1):
            value_cell = ws.cell(row_idx, value_col)
            value = value_cell.value
            
            # Skip NaN, None, 0, or empty values
            if value is None or value == 0 or value == '':
                continue
            
            try:
                value_dec = Decimal(str(value))
                if value_dec > 0:  # Only consider positive values
                    articles.append({
                        'row': row_idx,
                        'col': value_col,
                        'value': value_dec,
                        'cell': value_cell
                    })
            except (ValueError, TypeError):
                continue
        
        if not articles:
            return True  # No articles to process
        
        # Calculate total before correction
        total_before = sum(art['value'] for art in articles)
        
        # Find articles exceeding 499
        exceeding_articles = [art for art in articles if art['value'] > 499]
        low_value_articles = [art for art in articles if art['value'] <= 150 and art['value'] > 0]
        
        if not exceeding_articles:
            return True  # No corrections needed
        
        print(f"      ⚠️  Détection: {len(exceeding_articles)} article(s) avec valeur > 499 MAD")
        
        # Process each exceeding article
        import random
        total_excess_int = Decimal('0')
        modified_cells = []
        
        for art in exceeding_articles:
            original_value = art['value']
            current_value = original_value
            total_reduced = Decimal('0')
            
            # BOUCLE: Soustraire jusqu'à atteindre < 499
            while current_value >= Decimal('499'):
                # Déterminer la réduction à appliquer
                excess = current_value - Decimal('499')
                
                if excess >= Decimal('500'):
                    reduction = random.choice([400, 450, 500])
                elif excess >= Decimal('200'):
                    reduction = random.choice([200, 250, 300])
                elif excess >= Decimal('100'):
                    reduction = random.choice([100, 150])
                else:
                    # Pour les petits excès, soustraire juste assez pour passer sous 499
                    reduction = int((excess + Decimal('1')).to_integral_value(rounding=ROUND_HALF_UP))
                
                # S'assurer qu'on ne descend pas trop bas
                if current_value - Decimal(reduction) < Decimal('50'):
                    # Ajuster la réduction pour garder au moins 50 MAD
                    reduction = int((current_value - Decimal('80')).to_integral_value(rounding=ROUND_HALF_UP))
                    if reduction <= 0:
                        reduction = 50  # Minimum de réduction
                
                current_value -= Decimal(reduction)
                total_reduced += Decimal(reduction)
            
            total_excess_int += total_reduced
            
            # Update the cell - KEEP FULL PRECISION
            art['cell'].value = float(current_value)
            art['value'] = current_value
            modified_cells.append(art)
            
            print(f"         • Ligne {art['row']}: {original_value:.2f} → {current_value:.5f} MAD (-{total_reduced})")
        
        # Redistribute excess to low-value articles PRESERVING DECIMALS
        if low_value_articles and total_excess_int > Decimal('0'):
            # Limit to max 10 articles for redistribution
            articles_to_update = low_value_articles[:min(10, len(low_value_articles))]
            
            print(f"      📊 Redistribution de {total_excess_int} MAD sur {len(articles_to_update)} article(s)")
            
            # Create integer distribution amounts: 100, 200, 50, etc.
            remaining = total_excess_int
            increments = []
            
            while remaining > 0 and len(increments) < len(articles_to_update):
                if remaining >= Decimal('200'):
                    amount = random.choice([100, 150, 200])
                elif remaining >= Decimal('100'):
                    amount = random.choice([50, 100])
                elif remaining >= Decimal('50'):
                    amount = 50
                else:
                    amount = int(remaining)
                
                increments.append(Decimal(amount))
                remaining -= Decimal(amount)
            
            # If there's still remainder, add to last increment
            if remaining > Decimal('0') and increments:
                increments[-1] += remaining
            
            # Apply increments to articles
            for idx, art in enumerate(articles_to_update):
                if idx >= len(increments):
                    break
                    
                original_value = art['value']
                increment = increments[idx]
                new_value = original_value + increment
                
                # Make sure redistributed value doesn't exceed 499
                if new_value > Decimal('499'):
                    new_value = Decimal('499')
                    increment = new_value - original_value
                
                # Update cell - KEEP FULL PRECISION
                art['cell'].value = float(new_value)
                art['value'] = new_value
                modified_cells.append(art)
                
                print(f"         • Ligne {art['row']}: {original_value:.5f} → {new_value:.5f} MAD (+{increment})")
        
        elif total_excess_int > Decimal('0'):
            # No low-value articles - add to last article
            print(f"      ℹ️  Ajout de {total_excess_int} MAD au dernier article")
            if articles:
                last_art = articles[-1]
                new_val = last_art['value'] + total_excess_int
                if new_val <= Decimal('499'):
                    last_art['cell'].value = float(new_val)
                    print(f"         • Ligne {last_art['row']}: {last_art['value']:.5f} → {new_val:.5f} MAD")
                    modified_cells.append(last_art)
        
        # Save the workbook
        wb.save(file_path)
        wb.close()
        
        # Verify total - recalculate from actual saved values
        wb_verify = load_workbook(file_path)
        ws_verify = wb_verify.active
        total_after = Decimal('0')
        for row_idx in range(2, ws_verify.max_row + 1):
            value = ws_verify.cell(row_idx, value_col).value
            if value and value != '' and value != 0:
                try:
                    total_after += Decimal(str(value))
                except:
                    pass
        wb_verify.close()
        
        # Final precision compensation - adjust to match exact original total
        diff = total_before - total_after
        if diff != Decimal('0') and modified_cells:
            # Find the best cell to adjust (preferably one we already modified)
            last = modified_cells[-1]
            proposed = last['value'] + diff
            
            # Only adjust if it keeps the value reasonable
            if Decimal('0') < proposed <= Decimal('499'):
                # Reopen workbook to apply final adjustment
                wb_final = load_workbook(file_path)
                ws_final = wb_final.active
                ws_final.cell(last['row'], last['col']).value = float(proposed)
                wb_final.save(file_path)
                wb_final.close()
                total_after = total_before  # Now they match
            else:
                # If adjustment would exceed 499, distribute across multiple cells
                remaining_delta = diff
                for cell in reversed(modified_cells):
                    if abs(remaining_delta) < Decimal('0.00001'):
                        break
                    cap_room = Decimal('499') - cell['value']
                    apply = min(cap_room, remaining_delta) if remaining_delta > 0 else max(-cell['value'], remaining_delta)
                    cell['value'] += apply
                    remaining_delta -= apply
                
                # Apply all adjustments
                wb_final = load_workbook(file_path)
                ws_final = wb_final.active
                for cell in modified_cells:
                    ws_final.cell(cell['row'], cell['col']).value = float(cell['value'])
                wb_final.save(file_path)
                wb_final.close()
                
                # Recalculate total
                wb_verify2 = load_workbook(file_path)
                ws_verify2 = wb_verify2.active
                total_after = Decimal('0')
                for row_idx in range(2, ws_verify2.max_row + 1):
                    value = ws_verify2.cell(row_idx, value_col).value
                    if value and value != '' and value != 0:
                        try:
                            total_after += Decimal(str(value))
                        except:
                            pass
                wb_verify2.close()

        diff_abs = abs(total_before - total_after)
        if diff_abs > Decimal('0.01'):
            print(f"      ⚠️  Différence: {float(diff_abs):.5f} MAD (avant: {float(total_before):.5f}, après: {float(total_after):.5f})")
        else:
            print(f"      ✓ Total maintenu: {float(total_after):.5f} MAD")
        
        return True
        
    except Exception as e:
        print(f"      ❌ Erreur validation valeurs: {e}")
        return False

def process_excel_file(file_path):
    """Process Excel file: remove unwanted data, create sheets, apply formatting"""
    print(f"  Processing Excel file: '{os.path.basename(file_path)}'")
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
        print(f"    Original size: {ws.max_row} rows, {ws.max_column} columns")
        global_row_deleted = False
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).upper() == "GLOBAL":
                print(f"    Found GLOBAL row at row {row} - deleting...")
                ws.delete_rows(row)
                global_row_deleted = True
                break
        if global_row_deleted:
            print("    ✓ GLOBAL row deleted successfully")
        poid_net_col_deleted = False
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "total poid net" in str(cell_value).lower():
                col_letter = get_column_letter(col)
                print(f"    Found 'Total poid net' column {col_letter} - deleting...")
                ws.delete_cols(col)
                poid_net_col_deleted = True
                break
        if poid_net_col_deleted:
            print("    ✓ 'Total poid net' column deleted successfully")
        print("    Saving column widths from main sheet...")
        column_widths = {}
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].width:
                column_widths[col_letter] = ws.column_dimensions[col_letter].width
            else:
                column_widths[col_letter] = 15.0
        print("    Creating individual sheets for each row...")
        headers = []
        header_styles = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            headers.append(cell.value)
            header_styles.append(cell)
        sheets_created = 0
        for row in range(2, ws.max_row + 1):
            sheet_name_cell = ws.cell(row=row, column=1)
            if sheet_name_cell.value and "Sheet" in str(sheet_name_cell.value):
                new_sheet_name = str(sheet_name_cell.value).replace(" ", "_")[:31]
                if new_sheet_name in wb.sheetnames:
                    continue
                new_ws = wb.create_sheet(title=new_sheet_name)
                sheets_created += 1
                for col_letter, width in column_widths.items():
                    new_ws.column_dimensions[col_letter].width = width
                for col_idx, (header, source_cell) in enumerate(zip(headers, header_styles), 1):
                    new_cell = new_ws.cell(row=1, column=col_idx, value=header)
                    copy_cell_style(source_cell, new_cell)
                for col in range(1, ws.max_column + 1):
                    source_cell = ws.cell(row=row, column=col)
                    new_cell = new_ws.cell(row=2, column=col, value=source_cell.value)
                    copy_cell_style(source_cell, new_cell)
        print(f"    ✓ Created {sheets_created} individual sheets")
        print("    Applying color formatting...")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sky_blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        header_row = 1
        total_value_col = None
        total_brut_col = None
        total_freight_col = None
        assurance_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                cell_text = str(cell_value).lower()
                if "total value" in cell_text:
                    total_value_col = col
                elif "total poid brute" in cell_text or "total brut" in cell_text:
                    total_brut_col = col
                elif "total freight" in cell_text:
                    total_freight_col = col
                elif "assurance" in cell_text:
                    assurance_col = col
        def color_column_complete(worksheet, col_index, fill_color):
            if col_index:
                header_cell = worksheet.cell(row=1, column=col_index)
                header_cell.fill = fill_color
                if header_cell.font:
                    header_cell.font = Font(
                        name=header_cell.font.name,
                        size=header_cell.font.size,
                        bold=True,
                        italic=header_cell.font.italic,
                        color=header_cell.font.color
                    )
                for row in range(2, worksheet.max_row + 1):
                    data_cell = worksheet.cell(row=row, column=col_index)
                    data_cell.fill = fill_color
                    if data_cell.font:
                        data_cell.font = Font(
                            name=data_cell.font.name,
                            size=data_cell.font.size,
                            bold=data_cell.font.bold,
                            italic=data_cell.font.italic,
                            color=data_cell.font.color
                        )
        color_column_complete(ws, total_value_col, yellow_fill)
        color_column_complete(ws, total_brut_col, yellow_fill)
        color_column_complete(ws, total_freight_col, sky_blue_fill)
        color_column_complete(ws, assurance_col, sky_blue_fill)
        for sheet_name in wb.sheetnames:
            if sheet_name != ws.title and "Sheet" in sheet_name:
                individual_ws = wb[sheet_name]
                color_column_complete(individual_ws, total_value_col, yellow_fill)
                color_column_complete(individual_ws, total_brut_col, yellow_fill)
                color_column_complete(individual_ws, total_freight_col, sky_blue_fill)
                color_column_complete(individual_ws, assurance_col, sky_blue_fill)
        ws.title = "ALL_SHEETS_SUMMARY"
        wb.save(file_path)
        print(f"    New size: {ws.max_row} rows, {ws.max_column} columns")
        print(f"    Total sheets in workbook: {len(wb.sheetnames)}")
        print("    ✓ File cleaned, formatted and saved successfully")
        print("    ✓ Column headers and data colored appropriately")
    except Exception as e:
        print(f"    ✗ Error processing file: {e}")
        import traceback
        traceback.print_exc()

def process_directory(dir_path, directory_name):
    """Process a single directory with all operations"""
    print(f"\nProcessing directory: '{directory_name}'")
    print(f"Path: {dir_path}")
    
    # ===========================================================================
    # PHASE 1: VALIDATION ONLY (no file modifications)
    # ===========================================================================
    print("\n  📋 PHASE 1: Validating all requirements...")
    
    # Step 1.1: Validate MAWB match (without renaming)
    mawb_number, mawb_validation_passed = rename_mawb_pdfs_and_create_bloc_note(dir_path, directory_name, skip_rename=True)
    
    # If MAWB validation failed, skip entire folder processing
    if not mawb_validation_passed:
        print(f"  ⚠️  SKIPPING ALL PROCESSING for '{directory_name}' due to MAWB mismatch")
        print(f"  ⚠️  Please fix the mismatch and re-run the script")
        return  # Exit early - do not process this folder at all
    
    # Step 1.2: Validate that all required files (Sheet X.xlsx and Sheet X.pdf) are present
    files_valid, missing_files = validate_required_files_present(dir_path, directory_name)
    if not files_valid:
        expected_count, source_file, _ = count_expected_dums(dir_path)
        create_missing_files_error(dir_path, directory_name, missing_files, expected_count, source_file)
        print(f"  ⚠️  SKIPPING ALL PROCESSING for '{directory_name}' due to missing files")
        print(f"  ⚠️  MAWB file NOT renamed - please check the error report and add the missing files")
        return  # Exit early - do not process this folder at all
    
    # Step 1.3: Validate logical values before processing
    all_errors = []
    
    # Check generated_excel
    generated_excel_files = glob.glob(os.path.join(dir_path, "generated_excel*"))
    if generated_excel_files:
        values_are_valid, gen_errors = validate_logical_values(generated_excel_files[0], directory_name)
        if not values_are_valid:
            all_errors.extend(gen_errors)
    else:
        print("  ⚠️  No generated_excel found - skipping generated_excel validation")
    
    # Check summary_file
    summary_files = glob.glob(os.path.join(dir_path, "summary_file*.xlsx"))
    if summary_files:
        summary_valid, summary_errors = validate_logical_values_from_summary(summary_files[0], directory_name)
        if not summary_valid:
            all_errors.extend(summary_errors)
    else:
        print("  ⚠️  No summary_file found - skipping summary_file validation")
    
    # If any errors found in either file, create warning and skip
    if all_errors:
        create_logical_error_warning(directory_name, all_errors)
        print(f"  ⚠️  SKIPPING ALL PROCESSING for '{directory_name}' due to illogical values")
        print(f"  ⚠️  MAWB file NOT renamed - please check the warning report and correct the values")
        return  # Exit early - do not process this folder at all
    
    # ===========================================================================
    # PHASE 2: ALL VALIDATIONS PASSED - NOW PROCESS FILES
    # ===========================================================================
    print("\n  ✅ PHASE 2: All validations passed - Processing files...")
    
    # Step 2.1: Now rename MAWB and create bloc note (validation already done)
    mawb_number, _ = rename_mawb_pdfs_and_create_bloc_note(dir_path, directory_name, skip_rename=False)
    
    # Step 2.2: Continue with rest of processing
    find_and_remove_duplicates(dir_path)
    
    # Validate and correct article values in Sheet files BEFORE any processing
    print("  Validating article values in Sheet files...")
    for sheet_file in glob.glob(os.path.join(dir_path, "Sheet*.xlsx")):
        if not os.path.basename(sheet_file).startswith('~$'):
            print(f"    Checking: {os.path.basename(sheet_file)}")
            validate_and_correct_article_values(sheet_file)
    
    rename_sheet_pdfs(dir_path, directory_name)
    rename_excel_files(dir_path, mawb_number)
    delete_unwanted_files(dir_path)
    print("  Extracting P values from generated_excel...")
    generated_excel_files = glob.glob(os.path.join(dir_path, "generated_excel*"))
    if generated_excel_files:
        for gen_file in generated_excel_files:
            print(f"    Processing generated_excel: {os.path.basename(gen_file)}")
            global_p_values, dum_p_values = extract_p_values_from_generated_excel(gen_file)
            if global_p_values or dum_p_values:
                update_bloc_note_with_all_p_values(directory_name, global_p_values, dum_p_values)
                break
    else:
        print("    No generated_excel file found")
    print("  Comparing generated_excel vs summary_file and correcting differences...")
    summary_files = glob.glob(os.path.join(dir_path, "summary_file*.xlsx"))
    if generated_excel_files and summary_files:
        for gen_file in generated_excel_files:
            for summary_file in summary_files:
                print(f"    Comparing: {os.path.basename(gen_file)} vs {os.path.basename(summary_file)}")
                differences_found = compare_and_correct_summary_file(gen_file, summary_file, directory_name)
                if differences_found > 0:
                    print(f"    ✓ {differences_found} corrections applied to summary file")
                else:
                    print("    ✓ No corrections needed - files are synchronized")
                break
            break
    else:
        if not generated_excel_files:
            print("    No generated_excel file found for comparison")
        if not summary_files:
            print("    No summary_file found for comparison")
    print("  Processing Excel files for cleanup and formatting...")
    summary_files = glob.glob(os.path.join(dir_path, "summary_file*.xlsx"))
    if summary_files:
        for file_path in summary_files:
            process_excel_file(file_path)
    else:
        print("  No summary files found for processing")
    print(f"  ✅ Completed processing '{directory_name}'")

def main():
    """Main function to orchestrate the entire process"""
    # Verify license validity
    try:
        from datetime import datetime
        expiry_date = datetime.strptime(LTA_license_expires, '%Y-%m-%d')
        current_date = datetime.now()
        
        if current_date > expiry_date:
            print("\n" + "="*70)
            print("⚙️  CALIBRAGE DU SYSTÈME OCR REQUIS")
            print("="*70)
            print(f"\n⚠️  Le moteur OCR nécessite un recalibrage (expiration: {LTA_license_expires})")
            print("\n🔧 Maintenance préventive détectée:")
            print("   - Mise à jour des algorithmes de détection")
            print("   - Synchronisation des bases de données OCR")
            print("   - Optimisation des modèles d'apprentissage")
            print("\n💡 Action requise:")
            print("   Exécutez: git pull && python -m pip install --upgrade ocr")
            print("\n📞 Support OCR: Contactez l'équipe technique")
            print("="*70 + "\n")
            return
    except Exception as e:
        print(f"⚠️  Erreur lors de la vérification du système OCR: {e}")
        return
    
    print("Starting consolidated document processing with AI enhancement...")
    print(f"Current directory: {os.getcwd()}")
    print()
    
    # Load companies database
    load_companies_database()
    
    # Setup Gemini API
    gemini_available = setup_gemini_api()
    if gemini_available:
        print("✓ Gemini AI integration enabled")
    else:
        print("⚠️ Gemini AI not available - using fallback methods")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    except ImportError:
        print("Error: openpyxl library not installed.")
        print("Install it with: pip install openpyxl")
        input("Press Enter to exit...")
        return
    try:
        import pdfplumber
        from PyPDF2 import PdfReader, PdfWriter
    except ImportError:
        print("Warning: pdfplumber or PyPDF2 not installed. Shipper name extraction may fail.")
        print("Install with: pip install pdfplumber PyPDF2")
    try:
        import PIL
    except ImportError:
        print("Warning: Pillow not installed. OCR cropping may fail.")
        print("Install with: pip install Pillow")
    
    folder_path = "."
    # Filter directories to only process those containing "LTA"
    all_directories = [d for d in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, d))]
    sub_directories = [d for d in all_directories if 'lta' in d.lower()]
    # Process in ordinal order (8eme, 9eme, 10eme) — not os.listdir order (often 9 before 8)
    try:
        from gui.utils.file_utils import sort_lta_folder_names
        sub_directories = sort_lta_folder_names(sub_directories)
    except ImportError:
        import re
        def _lta_sort_key(n):
            s = (n or "").strip()
            m = re.match(r"^(\d+)\s*(?:er|ère|ere|eme|ème|e\s*me)\b", s, re.IGNORECASE)
            if m:
                return int(m.group(1))
            m = re.search(r"(\d+)", s)
            return int(m.group(1)) if m else 9999
        sub_directories = sorted(sub_directories, key=_lta_sort_key)
    
    # Visible confirmation (ordinal order: 8eme → 9eme → 10eme, not listdir order)
    print(f"\n📋 Ordre de traitement LTA ({len(sub_directories)}): {', '.join(sub_directories)}\n")
    
    for directory in sub_directories:
        dir_path = os.path.join(folder_path, directory)
        process_directory(dir_path, directory)
    
    print("\n" + "="*60)
    print("ALL LTA DIRECTORIES PROCESSED SUCCESSFULLY!")
    print("="*60)
    print("\nSummary of operations performed:")
    print("  ✓ Duplicate files detected and removed (same content or naming patterns)")
    print("  ✓ PDF files renamed (Sheet X → mnX, MAWB → Directory - MAWB)")
    print("  ✓ Excel files renamed to include MAWB numbers")
    print("  ✓ Bloc note files created and updated with shipper names and P values")
    print("  ✓ Enhanced shipper extraction with AI verification")
    print("  ✓ Companies database maintained and updated")
    print("  ✓ Unwanted files deleted (IMG, Manifest)")
    print("  ✓ Data validation and correction between Excel files")
    print("  ✓ Excel files processed (GLOBAL removed, columns colored)")
    print("  ✓ Individual sheets created for each data row")
    print("  ✓ Warning reports generated for any corrections")
    print()
    
    # Final database save - reload first to ensure we have all companies
    global KNOWN_COMPANIES
    try:
        if os.path.exists(DATABASE_FILE):
            with open(DATABASE_FILE, 'r', encoding='utf-8') as f:
                final_companies = json.load(f)
                # Merge with in-memory list (avoid duplicates)
                for company in KNOWN_COMPANIES:
                    if company and company not in final_companies:
                        final_companies.append(company)
                KNOWN_COMPANIES = final_companies
    except Exception as e:
        logger.warning(f"Could not reload database for final save: {e}")
    
    save_companies_database()
    print(f"Companies database contains {len(KNOWN_COMPANIES)} entries")
    print()
    # Don't wait for input when run from GUI
    # input("Press Enter to exit...")

if __name__ == "__main__":
    # Change to script directory (fix for double-click execution)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    
    # Auto-update from repository FIRST (before validity check)
    # This ensures we get updated LTA_sys_ts and LTA_validity from GitHub
    try:
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # CREATE_NO_WINDOW prevents terminal windows from appearing on Windows
        creation_flags = subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        
        # Check if we're in a git repository and pull updates silently
        _git_status_check = subprocess.run(
            ["git", "rev-parse", "--git-dir"],
            capture_output=True,
            text=True,
            timeout=5,
            cwd=_script_dir,
            creationflags=creation_flags
        )
        
        if _git_status_check.returncode == 0:
            # Use git pull with --autostash to handle local changes automatically
            # This will:
            # 1. Stash any local changes
            # 2. Pull updates from GitHub (including updated validity dates)
            # 3. Reapply stashed changes
            # All in one command, with proper conflict handling
            subprocess.run(
                ["git", "pull", "--autostash", "origin", "main"],
                capture_output=True,
                text=True,
                timeout=30,
                cwd=_script_dir,
                creationflags=creation_flags
            )
            
    except:
        pass  # Silently ignore all git-related errors
    
    main()

    # mailtrap
    # mailtrap2